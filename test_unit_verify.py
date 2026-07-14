#!/usr/bin/env python3

"""Unit-level verification for sync_register.py (no credentials required)."""



from __future__ import annotations



import os

import re

import sys

import tempfile

from datetime import datetime

from pathlib import Path



import sync_register as sr



SCRIPT_DIR = Path(__file__).resolve().parent

PASS = 0

FAIL = 0





def check(name: str, ok: bool, detail: str = "") -> None:

    global PASS, FAIL

    status = "PASS" if ok else "FAIL"

    if ok:

        PASS += 1

    else:

        FAIL += 1

    line = f"[{status}] {name}"

    if detail:

        line += f" — {detail}"

    print(line)





def main() -> int:

    print("=" * 60)

    print("sync_register.py 單元驗證")

    print("=" * 60)



    # 1. No issue link API in source

    src = (SCRIPT_DIR / "sync_register.py").read_text(encoding="utf-8")

    patterns = ["issueLink", "/issueLink", "issuelink", "已連結工作項目"]

    found = [p for p in patterns if p.lower() in src.lower()]

    check("4. 無 Jira issue link 程式碼", not found, "未找到 issueLink API" if not found else str(found))
    check(
        "4. 使用 remotelink Web 連結 API",
        "/remotelink" in src and "ensure_confluence_web_link" in src,
    )
    silent_client = sr.AtlassianClient(
        "https://example.atlassian.net",
        "cloud",
        "a@b.c",
        "token",
        silent_mode=True,
    )
    loud_client = sr.AtlassianClient(
        "https://example.atlassian.net",
        "cloud",
        "a@b.c",
        "token",
        silent_mode=False,
    )
    check(
        "寂靜寫入：Jira 路徑加 notifyUsers=false",
        silent_client._jira_path("/issue/PMWC-1", mutate=True)
        == "/issue/PMWC-1?notifyUsers=false"
        and silent_client._jira_path("/issue/X?deleteSubtasks=false", mutate=True)
        == "/issue/X?deleteSubtasks=false&notifyUsers=false"
        and loud_client._jira_path("/issue/PMWC-1", mutate=True) == "/issue/PMWC-1",
    )
    check(
        "寂靜寫入：Confluence 使用 minorEdit",
        "minorEdit" in src and "notifyUsers=false" in src,
    )
    check(
        "JQL 任務類型別名",
        sr.jql_issuetype_name_candidates("任務") == ["任務", "Task"]
        and sr.jql_issuetype_equals("Task") == "issuetype = Task",
    )
    check(
        "Register ID 允許 EVT2-01",
        bool(sr.REGISTER_ID_RE.match("EVT2-01"))
        and bool(sr.REGISTER_ID_RE.match("RF-01"))
        and sr.parse_register_id_from_summary(
            "EVT2-01_DOE process assumptions - QSI confirmation"
        )
        == "EVT2-01",
    )



    # 2. C table column structure

    check(

        "C 表欄位 = S 欄位 + LINK",

        sr.C_COLUMNS == sr.S_COLUMNS + [sr.LINK_COLUMN],

        str(sr.C_COLUMNS),

    )

    check(

        "LINK 為最後一欄",

        sr.C_COLUMNS[-1] == "LINK" and "LINK" not in sr.S_COLUMNS,

    )



    # 3. Status mapping — S/C 表預期為英文 Status
    cases = [
        ("Done", "完成"),
        ("DONE", "完成"),
        ("done", "完成"),
        ("Closed", "完成"),
        ("CLOSED", "完成"),
        ("closed", "完成"),
        ("Completed", "完成"),
        ("COMPLETED", "完成"),
        ("Closed - superseded", "完成"),
        ("Task done by vendor", "完成"),
        ("In progress", None),
        ("Open", None),
        ("Blocked", None),
        ("BLOCKED", None),
        ("Todo", None),
        ("Waiting", None),
        ("", None),
        ("Complete", None),  # 不含 Completed／Done／Closed
    ]
    status_ok = True
    for raw, expected in cases:
        got = sr.map_status_to_jira(raw)
        if got != expected:
            status_ok = False
            print(f"       {raw!r} -> {got} (期望 {expected})")
    check("6. Status 僅 Done/Closed/Completed→完成，其餘不變更", status_ok)

    contain_cases = [
        ("closed - superseded", "完成"),
        ("CLOSED - SuperSeded", "完成"),
        ("work Completed yesterday", "完成"),
        ("Marked as Done", "完成"),
        ("In Progress", None),
        ("Open", None),
        ("Blocked by vendor", None),
        ("blocked then closed", "完成"),  # 含 closed → 完成
    ]
    contain_ok = all(sr.map_status_to_jira(raw) == exp for raw, exp in contain_cases)
    if not contain_ok:
        for raw, exp in contain_cases:
            got = sr.map_status_to_jira(raw)
            if got != exp:
                print(f"       {raw!r} -> {got} (期望 {exp})")
    check("6. Status 子字串含 Closed/Done/Completed", contain_ok)

    no_change_cases = [
        "In progress",
        "in-progress",
        "Open",
        "Blocked",
        "BLOCKED",
        "Todo",
        "To do",
        "Waiting",
        "CANDIDATE",
        "RESUME",
        "ABORT",
        "Cancelled",
        "",
        "   ",
        "Complete",
        "进行中",
    ]
    no_change_ok = all(sr.map_status_to_jira(raw) is None for raw in no_change_cases)
    if not no_change_ok:
        for raw in no_change_cases:
            got = sr.map_status_to_jira(raw)
            if got is not None:
                print(f"       {raw!r} -> {got} (期望 None／不變更)")
    check("6. Status 其他值→不變更 Jira（None）", no_change_ok)




    # 4. Timeline fields (item 7)

    cfg = {"jira": {"start_date_field": "customfield_10015", "due_date_field": "duedate"}}

    sync_date = "2026-07-13"

    row = sr.RegisterRow(
        register_id="RF-01",
        workstream="RF",
        title="Test",
        opened="2026-06-25",
        target_close="2026-07-30",
    )
    tl = sr.build_timeline_fields(row, sync_date, cfg)
    check(
        "7. Opened+Target close → Start+Due",
        tl == {"duedate": "2026-07-30", "customfield_10015": "2026-06-25"},
        str(tl),
    )

    row_cn = sr.RegisterRow(
        register_id="RF-16",
        opened="2026-06-25",
        target_close="7月17日",
    )
    tl_cn = sr.build_timeline_fields(row_cn, sync_date, cfg)
    check(
        "7. 中文 Target close（7月17日）",
        tl_cn.get("customfield_10015") == "2026-06-25"
        and tl_cn.get("duedate") == f"{sync_date[:4]}-07-17",
        str(tl_cn),
    )

    year = datetime.now().year
    parse_cases = [
        ("7月31日", f"{year}-07-31"),
        ("7 月 31 日", f"{year}-07-31"),
        ("2026年7月31日", "2026-07-31"),
        ("2026年07月31号", "2026-07-31"),
        ("7/31/2026", "2026-07-31"),
        ("7/31/26", "2026-07-31"),
        ("7/31", f"{year}-07-31"),
        ("2026-07-31", "2026-07-31"),
        ("2026.07.31", "2026-07-31"),
        ("2026-07-31 00:00:00", "2026-07-31"),
        ("2026-07-31T12:00:00Z", "2026-07-31"),
        ("31-Jul-2026", "2026-07-31"),
        ("Jul 31, 2026", "2026-07-31"),
    ]
    parse_ok = True
    for raw, expect in parse_cases:
        got = sr.parse_flexible_date(raw)
        if got != expect:
            parse_ok = False
            print(f"       parse fail: {raw!r} -> {got!r} (期望 {expect!r})")
    # Excel serial: 2026-07-31 ≈ 46204 (1899-12-30 epoch)
    serial = (datetime(2026, 7, 31) - datetime(1899, 12, 30)).days
    if sr.parse_flexible_date(serial) != "2026-07-31":
        parse_ok = False
        print(f"       parse fail: serial {serial} -> {sr.parse_flexible_date(serial)!r}")
    if sr.parse_flexible_date(str(serial)) != "2026-07-31":
        parse_ok = False
        print(f"       parse fail: serial str -> {sr.parse_flexible_date(str(serial))!r}")
    if sr.parse_flexible_date("7") is not None:
        parse_ok = False
        print("       parse fail: '7' should not be Excel serial")
    check("7. parse_flexible_date 中文/斜線/ISO/序號", parse_ok)

    check("7. 解析 7月17日", sr.parse_flexible_date("7月17日") == f"{year}-07-17")
    check("7. 解析 2026年7月17日", sr.parse_flexible_date("2026年7月17日") == "2026-07-17")

    # Opened only → Start only（不寫 Due、不要求 Target close）
    row_opened_only = sr.RegisterRow(
        register_id="RF-03",
        opened="2026-06-01",
        target_close="",
    )
    tl_opened = sr.build_timeline_fields(row_opened_only, sync_date, cfg)
    check(
        "7. 僅 Opened → 只設 Start",
        tl_opened == {"customfield_10015": "2026-06-01"},
        str(tl_opened),
    )

    # Target close only（無 Opened）→ 只設 Due（不填 Start）
    row_due_only = sr.RegisterRow(
        register_id="RF-04",
        opened="",
        target_close="2026-08-15",
    )
    tl_due = sr.build_timeline_fields(row_due_only, sync_date, cfg)
    check(
        "7. 僅 Target close → 只設 Due",
        tl_due == {"duedate": "2026-08-15"},
        str(tl_due),
    )

    # Opened 中文 + Target close 斜線
    row_mixed = sr.RegisterRow(
        register_id="RF-06",
        opened="6月25日",
        target_close="7/31/26",
    )
    tl_mixed = sr.build_timeline_fields(row_mixed, sync_date, cfg)
    check(
        "7. Opened 中文 + Target close 斜線",
        tl_mixed == {
            "customfield_10015": f"{year}-06-25",
            "duedate": "2026-07-31",
        },
        str(tl_mixed),
    )

    # due < start → year adjust
    row_year = sr.RegisterRow(
        register_id="RF-05",
        opened="2026-11-01",
        target_close="1/15",
    )
    tl_year = sr.build_timeline_fields(row_year, sync_date, cfg)
    check(
        "7. due < start 時順延隔年",
        tl_year.get("customfield_10015") == "2026-11-01"
        and tl_year.get("duedate") == "2027-01-15",
        str(tl_year),
    )

    adf = sr.text_to_adf("Register ID: RF-01")
    adf_ok = (
        adf["type"] == "doc"
        and adf["version"] == 1
        and adf["content"][0]["type"] == "paragraph"
        and adf["content"][0]["content"][0]["text"] == "Register ID: RF-01"
    )
    row_desc = sr.RegisterRow(register_id="RF-01", title="Foo", description="Bar")
    adf2 = sr.build_jira_description(row_desc)
    adf2_ok = len(adf2["content"]) == 3
    check("Jira description ADF 格式", adf_ok and adf2_ok)

    row_empty = sr.RegisterRow(register_id="RF-02", workstream="RF", title="No date")

    check(
        "7. Opened 與 Target close 皆空白不更新",
        sr.build_timeline_fields(row_empty, sync_date, cfg) == {},
    )

    # 空白／無法解析 → 不寫入對應 Jira 欄（omit，不清空）
    row_blank_opened = sr.RegisterRow(
        register_id="RF-10",
        opened="",
        target_close="2026-09-01",
    )
    tl_blank_opened = sr.build_timeline_fields(row_blank_opened, sync_date, cfg)
    check(
        "7. 空白 Opened → 無 Start（僅 Due）",
        tl_blank_opened == {"duedate": "2026-09-01"},
        str(tl_blank_opened),
    )

    row_blank_due = sr.RegisterRow(
        register_id="RF-11",
        opened="2026-05-01",
        target_close="",
    )
    tl_blank_due = sr.build_timeline_fields(row_blank_due, sync_date, cfg)
    check(
        "7. 空白 Target close → 無 Due（僅 Start）",
        tl_blank_due == {"customfield_10015": "2026-05-01"},
        str(tl_blank_due),
    )

    row_garbage = sr.RegisterRow(
        register_id="RF-12",
        opened="TBD",
        target_close="asap / next week",
    )
    check(
        "7. 無法解析文字 → 不更新 Start/Due",
        sr.build_timeline_fields(row_garbage, sync_date, cfg) == {},
    )

    row_garbage_opened = sr.RegisterRow(
        register_id="RF-13g",
        opened="N/A",
        target_close="2026-10-15",
    )
    tl_garbage_opened = sr.build_timeline_fields(row_garbage_opened, sync_date, cfg)
    check(
        "7. Opened 垃圾文字 + 有效 Due → 僅 Due",
        tl_garbage_opened == {"duedate": "2026-10-15"},
        str(tl_garbage_opened),
    )

    row_garbage_due = sr.RegisterRow(
        register_id="RF-14g",
        opened="2026-04-20",
        target_close="待定",
    )
    tl_garbage_due = sr.build_timeline_fields(row_garbage_due, sync_date, cfg)
    check(
        "7. 有效 Opened + Due 垃圾文字 → 僅 Start",
        tl_garbage_due == {"customfield_10015": "2026-04-20"},
        str(tl_garbage_due),
    )

    # 僅一邊可解析時不做 year+1（due 雖 < 當年度某日，但無 parsed start）
    row_due_only_early = sr.RegisterRow(
        register_id="RF-15g",
        opened="",
        target_close="1/15",
    )
    tl_due_early = sr.build_timeline_fields(row_due_only_early, sync_date, cfg)
    check(
        "7. 僅 Due 可解析時不做 year+1",
        tl_due_early == {"duedate": f"{year}-01-15"},
        str(tl_due_early),
    )



    # 5. HTML table — no blank rows, LINK in last column (items 2, 3)

    rows = [

        sr.RegisterRow(
            register_id="RF-13",
            workstream="EVT1 RF testing",
            title="Closed item",
            ball_with="VOX",
            priority="P1",
            status="Closed - superseded",
            jira_key="PMWC-100",
        ),
        sr.RegisterRow(
            register_id="RF-22",
            workstream="EVT1 RF testing",
            title="Another closed",
            ball_with="VOX",
            priority="P1",
            status="Closed - superseded",
            jira_key="PMWC-101",
        ),
        sr.RegisterRow(
            register_id="ECO-05",
            workstream="Layout ECO / re-gerber",
            title="Open item",
            ball_with="Joint",
            priority="P2",
            status="Open",
            jira_key="PMWC-102",
        ),

    ]

    site = "https://qsiaiot.atlassian.net"

    html = sr.build_confluence_html_table(rows, site)

    data_rows = html.count("<tr>") - 1

    blank_rows = bool(

        re.search(rf"<tr>\s*(<td><p></p></td>\s*){{{len(sr.C_COLUMNS)}}}</tr>", html)

    )

    links = re.findall(r'href="[^"]+/browse/(PMWC-\d+)"', html)

    ids = re.findall(r"<td><p>([A-Z]+-\d+)</p></td>", html)

    header_match = re.search(r"<tr>(.*?)</tr>", html, flags=re.DOTALL)

    headers = re.findall(r"<strong>([^<]+)</strong>", header_match.group(1) if header_match else "")



    check("2. HTML 無空白 spacer 列", not blank_rows, f"data_rows={data_rows}")

    source_html = sr.build_confluence_source_link_html(
        "https://example.com/register.xlsx", "S 表格測試"
    )
    page_html = sr.build_confluence_page_html(
        rows[:1],
        site,
        source_link_url="https://example.com/register.xlsx",
        source_link_title="S 表格測試",
    )
    check(
        "C 表頂部 S 表格連結",
        "S 表格：" in source_html
        and "https://example.com/register.xlsx" in source_html
        and page_html.startswith("<p><strong>S 表格：</strong>")
        and "<table" in page_html,
    )

    # 已移除頁首「立即同步」；即使誤傳舊參數也不應再出現紅按鈕
    check(
        "C 表頂部不再渲染「立即同步」連結",
        "立即同步" not in source_html
        and "background-color:#DE350B" not in source_html
        and "立即同步" not in page_html
        and "sync-register.yml" not in page_html,
    )
    corrupted_title = "C27 Open Issues Register (S ??)????"
    fixed_url, fixed_title = sr.resolve_sharepoint_source_link(
        {"confluence": {"source_link_title": corrupted_title, "source_link_url": "https://ex"}, "sharepoint": {}}
    )
    fixed_html = sr.build_confluence_source_link_html("https://ex", corrupted_title)
    check(
        "source_link_title 含 ?? 時改用 UTF-8 預設中文標題",
        fixed_url == "https://ex"
        and fixed_title == sr.DEFAULT_SOURCE_LINK_TITLE
        and "??" not in fixed_title
        and "表格" in fixed_title
        and "??" not in fixed_html
        and "S 表格" in fixed_html
        and sr.DEFAULT_SOURCE_LINK_TITLE in fixed_html,
    )

    check(

        "3. 每列 LINK 欄有 PMWC 連結",

        len(links) == len(rows),

        f"links={links}",

    )

    check(

        "3. Closed 項目 RF-13/RF-22 含連結",

        "PMWC-100" in links and "PMWC-101" in links,

    )

    check(

        "LINK 欄為表頭最後一欄",

        headers[-1] == "LINK" and headers[1] == "Workstream",

        f"headers={headers}",

    )



    # 6. S master — C row count equals S (item 1, logic)

    check(

        "1. C 列數 = S 列數（邏輯）",

        data_rows == len(rows) and set(ids) == {"RF-13", "RF-22", "ECO-05"},

        f"ids={ids}",

    )



    # Example HTML row snippet

    rf13_html = re.search(r"<tr>.*?RF-13.*?</tr>", html, flags=re.DOTALL)

    if rf13_html:

        print(f"\n  範例 HTML 列（RF-13）:\n  {rf13_html.group(0)[:300]}...")



    # 7. S row mapping — 直接對應 S 表表頭

    s_dict = {
        "ID": "RF-01",
        "Workstream": "EVT1 RF testing",
        "Title": "BT LE-2M Rx sensitivity failing",
        "Description / current state": "QSI to investigate LE-2M Rx path",
        "Next action": "QSI to investigate LE-2M Rx path; run RF-03 (XO trim) check first before treating as hardware.",
        "Ball with": "QSI",
        "Priority": "P1",
        "Status": "Open",
        "Opened": "2026-06-22",
    }
    reg = sr.s_row_to_register(s_dict)
    check(
        "S 欄位直接對應",
        reg is not None
        and reg.register_id == "RF-01"
        and reg.workstream == "EVT1 RF testing"
        and reg.title == "BT LE-2M Rx sensitivity failing"
        and reg.description == "QSI to investigate LE-2M Rx path"
        and reg.next_action.startswith("QSI to investigate")
        and reg.ball_with == "QSI"
        and reg.priority == "P1"
        and reg.status == "Open"
        and reg.opened == "2026-06-22",
        f"ball={reg.ball_with if reg else None!r}, pri={reg.priority if reg else None!r}, st={reg.status if reg else None!r}",
    )
    check(
        "Jira Epic 讀 Priority 欄",
        reg is not None and sr.jira_epic_priority(reg) == "P1",
    )
    check(
        "Jira Status 讀 Status 欄（Open→不變更）",
        reg is not None and sr.map_status_to_jira(sr.jira_status_source(reg)) is None,
    )

    check(
        "Jira summary = ID_Title",
        sr.jira_summary(
            "RF-05", "BT Rx max-input-level sweep to +10 dBm"
        )
        == "RF-05_BT Rx max-input-level sweep to +10 dBm",
    )
    check(
        "Jira summary ID重複前綴",
        sr.jira_summary("RF-14", "Some title", duplicate=True)
        == "ID重複 RF-14_Some title"
        and sr.jira_summary("RF-14", "Some title", duplicate=False)
        == "RF-14_Some title",
    )
    check(
        "parse summary 含 ID重複前綴",
        sr.parse_register_id_from_summary("ID重複 RF-14_Some title") == "RF-14"
        and sr.parse_register_id_from_summary("ID重複_RF-14_Some title") == "RF-14"
        and sr.parse_register_id_from_summary("RF-14_Some title") == "RF-14",
    )
    dup_rows = [
        sr.RegisterRow(register_id="RF-01", title="A"),
        sr.RegisterRow(register_id="RF-02", title="B"),
        sr.RegisterRow(register_id="RF-01", title="A2"),
        sr.RegisterRow(register_id="RF-03", title="C"),
    ]
    dups = sr.find_duplicate_register_ids(dup_rows)
    check(
        "偵測 S 表重複 Register ID",
        dups == {"RF-01"},
        str(dups),
    )
    # 唯一後不應再加前綴（模擬下次 sync）
    unique_again = [
        sr.RegisterRow(register_id="RF-01", title="Only one"),
        sr.RegisterRow(register_id="RF-02", title="B"),
    ]
    check(
        "重複解決後 summary 無 ID重複前綴",
        not sr.find_duplicate_register_ids(unique_again)
        and sr.jira_summary(
            "RF-01",
            "Only one",
            duplicate="RF-01" in sr.find_duplicate_register_ids(unique_again),
        )
        == "RF-01_Only one",
    )

    rank_rows = [
        sr.RegisterRow(register_id="A-1", priority="P1", status="Closed"),
        sr.RegisterRow(register_id="A-2", priority="P1", status="Open"),
        sr.RegisterRow(
            register_id="A-3",
            priority="P1",
            status="In progress",
            opened="2026-06-01",
            target_close="2026-07-17",
        ),
        sr.RegisterRow(register_id="A-4", priority="P1", status="In progress"),
        sr.RegisterRow(
            register_id="A-5",
            priority="P1",
            status="Open",
            opened="2026-06-01",
            target_close="2026-07-31",
        ),
    ]
    ordered = sr.sort_rows_for_jira_rank(rank_rows)
    check(
        "Jira Rank 順序：時間軸→進行中→其他→完成",
        [r.register_id for r in ordered] == ["A-3", "A-5", "A-4", "A-2", "A-1"],
        str([r.register_id for r in ordered]),
    )



    # 8. Column letter mapping — sparse row

    from openpyxl import Workbook



    wb = Workbook()

    ws = wb.active

    ws.title = "Register"

    headers_excel = [
        "ID",
        "Workstream",
        "Title",
        "Description / current state",
        "Next action",
        "Ball with",
        "Priority",
        "Status",
        "Opened",
    ]
    for col_idx, header in enumerate(headers_excel, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    ws["A2"] = "RF-99"
    ws["B2"] = "EVT1 RF testing"
    ws["C2"] = "Title row"
    ws["D2"] = "Desc row"
    ws["E2"] = "Next action text"
    ws["G2"] = "P2"
    ws["H2"] = "Blocked"
    ws["I2"] = "2026-07-01"
    with tempfile.TemporaryDirectory() as tmp:
        sparse_xlsx = Path(tmp) / "sparse.xlsx"
        wb.save(sparse_xlsx)
        wb.close()
        _, sparse_rows = sr.load_register_rows_from_excel(sparse_xlsx, "Register")
        sparse = next((r for r in sparse_rows if r.register_id == "RF-99"), None)
        check(
            "欄位字母對應（稀疏列 F 空白）",
            sparse is not None
            and sparse.next_action == "Next action text"
            and sparse.ball_with == ""
            and sparse.priority == "P2"
            and sparse.status == "Blocked"
            and sparse.opened == "2026-07-01",
            f"ball={sparse.ball_with if sparse else None!r}, pri={sparse.priority if sparse else None!r}",
        )



    # 9. parse_jira_map — LINK 欄與舊版 JIRA 欄

    new_body = (

        '<table><tr><th>ID</th><th>JIRA</th><th>LINK</th></tr>'

        '<tr><td>RF-01</td><td>EVT1</td>'

        '<td><a href="https://qsiaiot.atlassian.net/browse/PMWC-121">PMWC-121</a></td></tr>'

        "</table>"

    )

    legacy_body = (

        "| ID | JIRA | Workstream |\n"

        "| --- | --- | --- |\n"

        "| RF-13 | [PMWC-100](https://qsiaiot.atlassian.net/browse/PMWC-100) | Title |\n"

    )

    new_map = sr.parse_jira_map_from_confluence_body(new_body)

    legacy_map = sr.parse_jira_map_from_confluence_body(legacy_body)

    check("LINK 欄解析 PMWC key", new_map.get("RF-01") == "PMWC-121", str(new_map))

    check("舊版 JIRA 欄相容", legacy_map.get("RF-13") == "PMWC-100", str(legacy_map))



    # 10. Priority epic parent logic exists (item 5, code inspection)

    has_parent = 'parent": {"key": epic_key}' in src or '"parent": {"key": epic_key}' in src

    check("5. Priority 掛 Epic parent 邏輯存在", has_parent)



    # 11. Register diff / snapshot

    old_rows = {

        "RF-01": {

            "jira": "RF",

            "workstream": "Old title",

            "title": "Desc",

            "description": "",

            "next_action": "",

            "ball_with": "QSI",
            "priority": "P1",
            "status": "Open",

            "opened": "",

            "target_close": "",

            "next_milestone": "",

            "source": "",

            "qsi_comment": "",

        },

    }

    new_rows = [

        sr.RegisterRow(

            register_id="RF-01",

            jira="RF",

            workstream="New title",
            ball_with="QSI",
            priority="P1",
            status="In progress",

        ),

        sr.RegisterRow(

            register_id="RF-03",

            jira="ECO",

            workstream="Added",
            priority="P1",
            status="Open",

        ),

    ]

    diff = sr.compute_register_diff(old_rows, new_rows, snapshot_saved_at="2026-07-10T10:00:00")

    check(

        "Diff 新增/移除/欄位",

        diff.added_ids == ["RF-03"]

        and diff.removed_ids == []

        and len(diff.status_changes) == 1

        and diff.status_changes[0].register_id == "RF-01"

        and any(c.field_label == "Workstream" for c in diff.field_changes),

        f"added={diff.added_ids}, removed={diff.removed_ids}",

    )



    with tempfile.TemporaryDirectory() as tmp:

        snap_path = Path(tmp) / "last_snapshot.json"

        sr.save_last_snapshot(snap_path, new_rows)

        loaded, saved_at = sr.load_last_snapshot(snap_path)

        check(

            "快照存取",

            "RF-03" in loaded and loaded["RF-03"]["workstream"] == "Added" and bool(saved_at),

        )




    # 11b. Diff report formatting / email body
    empty_diff = sr.RegisterDiff(snapshot_saved_at="2026-07-10T10:00:00")
    empty_report_text = sr.format_diff_report(empty_diff)
    check(
        "Diff report conclusion no change",
        "【結論】無差異" in empty_report_text,
        empty_report_text[-120:],
    )

    status_diff = sr.RegisterDiff(
        status_changes=[
            sr.FieldChange(
                register_id="RF-01",
                field_label="Status",
                old_value="Open",
                new_value="Closed",
            )
        ],
        snapshot_saved_at="2026-07-10T10:00:00",
    )
    status_report_text = sr.format_diff_report(status_diff)
    check(
        "Diff report shows status section",
        "【狀態變更】" in status_report_text
        and "RF-01" in status_report_text
        and "Open" in status_report_text
        and "Closed" in status_report_text
        and "【結論】有差異" in status_report_text,
        status_report_text,
    )

    # 新增列應帶 Title；長字串差異不可截成看起來「沒變」
    added_with_title = sr.RegisterDiff(
        added_ids=["TST-99"],
        id_titles={"TST-99": "Cursor mail diff test"},
        snapshot_saved_at="2026-07-10T10:00:00",
    )
    added_text = sr.format_diff_report(added_with_title)
    check(
        "Diff report 新增含 Title",
        "+ TST-99 — Cursor mail diff test" in added_text
        and "【結論】有差異" in added_text
        and "【欄位變更】0 筆" in added_text,
        added_text,
    )
    long_prefix = "X" * 90
    old_long = long_prefix + "BEFORE_TAIL"
    new_long = long_prefix + "AFTER_TAIL"
    old_fmt, new_fmt = sr._format_diff_value_pair(old_long, new_long)
    check(
        "Diff 長字串顯示移到差異點",
        old_fmt != new_fmt and "BEFORE_TAIL" in old_fmt and "AFTER_TAIL" in new_fmt,
        f"old={old_fmt!r} new={new_fmt!r}",
    )
    crlf_old = {
        "RF-01": {
            "jira": "",
            "workstream": "",
            "title": "T",
            "description": "line1\r\nline2",
            "next_action": "",
            "ball_with": "",
            "priority": "P1",
            "status": "Open",
            "opened": "",
            "target_close": "",
            "next_milestone": "",
            "source": "",
            "qsi_comment": "",
        }
    }
    crlf_new = [
        sr.RegisterRow(
            register_id="RF-01",
            title="T",
            description="line1\nline2",
            priority="P1",
            status="Open",
        )
    ]
    crlf_diff = sr.compute_register_diff(crlf_old, crlf_new, snapshot_saved_at="t")
    check(
        "Diff 正規化 CRLF 不報假差異",
        not crlf_diff.field_changes and not crlf_diff.status_changes,
        f"fields={crlf_diff.field_changes}",
    )

    email_body = sr.build_diff_email_body(
        empty_report_text,
        sr.SyncReport(),
        verdict="無差異",
        s_link_url="https://example.com/register.xlsx",
        s_link_title="SharePoint Register",
        confluence_url="https://confluence.example/page",
    )
    check(
        "Diff email body includes S link",
        "https://example.com/register.xlsx" in email_body
        and "SharePoint Register" in email_body,
        email_body[:200],
    )
    check(
        "Diff email body first line 無差異",
        email_body.startswith("【無差異】"),
        email_body.splitlines()[0] if email_body else "",
    )
    email_changed = sr.build_diff_email_body(
        "========== S 表格變更摘要 ==========\n【新增】1 筆",
        sr.SyncReport(),
        verdict="有差異",
    )
    html_changed = sr.build_diff_email_html_body(email_changed, verdict="有差異")
    check(
        "Diff email plain first line 有差異",
        email_changed.startswith("【有差異】"),
        email_changed.splitlines()[0],
    )
    check(
        "Diff email HTML first line red 有差異",
        'style="color:red;font-weight:bold">有差異</span>' in html_changed
        and "【有差異】" not in html_changed.split("<pre", 1)[0],
        html_changed[:280],
    )
    email_first = sr.build_diff_email_body("snap", sr.SyncReport(), verdict="首次快照")
    html_first = sr.build_diff_email_html_body(email_first, verdict="首次快照")
    check(
        "Diff email first-run plain 首次快照",
        email_first.startswith("【首次快照】"),
        email_first.splitlines()[0],
    )
    check(
        "Diff email first-run HTML not red 有差異",
        "首次快照" in html_first
        and "color:red" not in html_first
        and "有差異" not in html_first.split("<pre", 1)[0],
        html_first[:280],
    )
    first_diff = sr.RegisterDiff(is_first_run=True, added_ids=["A-1"])
    none_diff = sr.RegisterDiff(is_first_run=False)
    changed_diff = sr.RegisterDiff(is_first_run=False, added_ids=["B-1"])
    check(
        "resolve_diff_email_verdict mapping",
        sr.resolve_diff_email_verdict(first_diff) == "首次快照"
        and sr.resolve_diff_email_verdict(none_diff) == "無差異"
        and sr.resolve_diff_email_verdict(changed_diff) == "有差異",
    )

    # 11c. Mail backend selection / Graph payload / SMTP host inference
    check(
        "infer_smtp_host gmail",
        sr.infer_smtp_host("bob@gmail.com") == "smtp.gmail.com",
    )
    check(
        "infer_smtp_host googlemail",
        sr.infer_smtp_host("bob@googlemail.com") == "smtp.gmail.com",
    )
    check(
        "infer_smtp_host corp M365",
        sr.infer_smtp_host("bob.lin@qsitw.com") == "smtp.office365.com",
    )

    saved_env = {
        k: os.environ.pop(k, None)
        for k in (
            "MAIL_BACKEND",
            "GRAPH_TENANT_ID",
            "GRAPH_CLIENT_ID",
            "GRAPH_CLIENT_SECRET",
            "AZURE_TENANT_ID",
            "AZURE_CLIENT_ID",
            "AZURE_CLIENT_SECRET",
            "MAIL_USERNAME",
            "MAIL_PASSWORD",
            "SYNC_NOTIFY_EMAIL",
            "SYNC_NOTIFY_CC",
            "SMTP_HOST",
        )
    }
    try:
        check(
            "resolve_mail_backend auto→smtp without Graph",
            sr.resolve_mail_backend({"notify": {"mail_backend": "auto"}}) == "smtp",
        )
        os.environ["GRAPH_TENANT_ID"] = "tid"
        os.environ["GRAPH_CLIENT_ID"] = "cid"
        os.environ["GRAPH_CLIENT_SECRET"] = "sec"
        check(
            "resolve_mail_backend auto→graph when Graph present",
            sr.resolve_mail_backend({"notify": {}}) == "graph",
        )
        check(
            "resolve_mail_backend explicit smtp overrides Graph",
            sr.resolve_mail_backend({"notify": {"mail_backend": "smtp"}}) == "smtp",
        )

        to_list, cc_list = sr.resolve_notify_recipients(
            {
                "notify": {
                    "email_to": "bob.lin@qsitw.com",
                    "email_cc": "shannon.chang@qsitw.com",
                }
            }
        )
        check(
            "resolve_notify_recipients To/Cc split from config",
            to_list == ["bob.lin@qsitw.com"]
            and cc_list == ["shannon.chang@qsitw.com"],
            f"to={to_list} cc={cc_list}",
        )
        os.environ["SYNC_NOTIFY_EMAIL"] = "bob.lin@qsitw.com"
        os.environ["SYNC_NOTIFY_CC"] = "shannon.chang@qsitw.com"
        to_env, cc_env = sr.resolve_notify_recipients({"notify": {}})
        check(
            "resolve_notify_recipients from SYNC_NOTIFY_* env",
            to_env == ["bob.lin@qsitw.com"]
            and cc_env == ["shannon.chang@qsitw.com"],
            f"to={to_env} cc={cc_env}",
        )
        os.environ.pop("SYNC_NOTIFY_EMAIL", None)
        os.environ.pop("SYNC_NOTIFY_CC", None)

        payload = sr.build_graph_send_mail_payload(
            subject="Subj",
            body="Body text",
            to_addr="bob.lin@qsitw.com",
            cc_addrs="shannon.chang@qsitw.com",
        )
        check(
            "Graph sendMail payload shape with Cc",
            payload["message"]["subject"] == "Subj"
            and payload["message"]["body"]["contentType"] == "Text"
            and payload["message"]["body"]["content"] == "Body text"
            and payload["message"]["toRecipients"][0]["emailAddress"]["address"]
            == "bob.lin@qsitw.com"
            and len(payload["message"]["toRecipients"]) == 1
            and payload["message"]["ccRecipients"][0]["emailAddress"]["address"]
            == "shannon.chang@qsitw.com"
            and payload.get("saveToSentItems") is False,
            str(payload)[:240],
        )
        payload_html = sr.build_graph_send_mail_payload(
            subject="Subj",
            body="【有差異】\nplain",
            to_addr="bob.lin@qsitw.com",
            html_body='<span style="color:red;font-weight:bold">有差異</span>',
        )
        check(
            "Graph sendMail HTML contentType",
            payload_html["message"]["body"]["contentType"] == "HTML"
            and "color:red" in payload_html["message"]["body"]["content"]
            and "ccRecipients" not in payload_html["message"],
            str(payload_html["message"]["body"])[:200],
        )

        # 郵件主旨 [bracket] = S 表檔名（非 PMWC Sync）
        cd_plain = sr.parse_content_disposition_filename(
            'attachment; filename="C27_VOX-QSI_Open_Issues_Register_20260702.xlsx"'
        )
        cd_star = sr.parse_content_disposition_filename(
            "attachment; filename*=UTF-8''C27_VOX-QSI_Open_Issues_Register_20260702.xlsx"
        )
        check(
            "Content-Disposition filename= / filename*",
            cd_plain == "C27_VOX-QSI_Open_Issues_Register_20260702.xlsx"
            and cd_star == "C27_VOX-QSI_Open_Issues_Register_20260702.xlsx",
            f"plain={cd_plain!r} star={cd_star!r}",
        )
        token = sr.resolve_register_subject_token(
            "C27_VOX-QSI_Open_Issues_Register_20260702.xlsx", {}
        )
        subj = sr.build_diff_email_subject(
            subject_token=token,
            verdict="有差異",
            stamp="2026-07-14 11:00",
        )
        check(
            "diff email subject uses S filename bracket",
            token == "C27_VOX-QSI_Open_Issues_Register_20260702"
            and subj
            == "[C27_VOX-QSI_Open_Issues_Register_20260702] 有差異 — 2026-07-14 11:00"
            and "PMWC Sync" not in subj,
            subj,
        )
        fallback_token = sr.resolve_register_subject_token(
            "",
            {
                "sharepoint": {
                    "register_filename": "C27_VOX-QSI_Open_Issues_Register_20260702.xlsx"
                }
            },
        )
        check(
            "subject token falls back to sharepoint.register_filename",
            fallback_token == "C27_VOX-QSI_Open_Issues_Register_20260702",
            fallback_token,
        )
        skip_cache_token = sr.resolve_register_subject_token(
            "register_latest.xlsx",
            {"confluence": {"source_link_title": "C27 Open Issues Register (S 表格)"}},
        )
        check(
            "subject token skips register_latest cache name",
            skip_cache_token == "C27_Open_Issues_Register_(S_表格)",
            skip_cache_token,
        )

        # SMTP path with mock（預設走 smtp.gmail.com）
        import smtplib as _smtplib

        class _FakeSMTP:
            last = None

            def __init__(self, host, port, timeout=None):
                self.host = host
                self.port = port
                _FakeSMTP.last = self

            def __enter__(self):
                return self

            def __exit__(self, *args):
                return False

            def ehlo(self):
                return True

            def starttls(self):
                return True

            def login(self, user, password):
                self.user = user
                self.password = password

            def send_message(self, msg):
                self.msg = msg

        for k in (
            "GRAPH_TENANT_ID",
            "GRAPH_CLIENT_ID",
            "GRAPH_CLIENT_SECRET",
        ):
            os.environ.pop(k, None)
        saved_smtp = _smtplib.SMTP
        try:
            _smtplib.SMTP = _FakeSMTP  # type: ignore[misc,assignment]
            os.environ["MAIL_USERNAME"] = "bot@gmail.com"
            os.environ["MAIL_PASSWORD"] = "app-pass"
            os.environ["SYNC_NOTIFY_EMAIL"] = "bob.lin@qsitw.com"
            os.environ.pop("SMTP_HOST", None)
            os.environ.pop("SYNC_NOTIFY_CC", None)
            sr.send_diff_email(
                {
                    "notify": {
                        "mail_backend": "smtp",
                        "email_cc": "shannon.chang@qsitw.com",
                    }
                },
                subject="[PMWC Sync] test",
                body="【有差異】\nhello",
                html_body=(
                    '<span style="color:red;font-weight:bold">有差異</span>'
                    "<pre>hello</pre>"
                ),
            )
            fake = _FakeSMTP.last
            multipart_ok = False
            if fake is not None and getattr(fake, "msg", None) is not None:
                msg = fake.msg
                multipart_ok = (
                    msg.is_multipart()
                    and any(
                        p.get_content_type() == "text/plain" for p in msg.iter_parts()
                    )
                    and any(
                        p.get_content_type() == "text/html" for p in msg.iter_parts()
                    )
                )
            check(
                "send_diff_email SMTP uses inferred Gmail host",
                fake is not None
                and fake.host == "smtp.gmail.com"
                and fake.port == 587
                and fake.user == "bot@gmail.com"
                and fake.msg["To"] == "bob.lin@qsitw.com",
                f"host={getattr(fake, 'host', None)}",
            )
            check(
                "send_diff_email SMTP sets Cc header separately",
                fake is not None
                and fake.msg["Cc"] == "shannon.chang@qsitw.com"
                and fake.msg["To"] == "bob.lin@qsitw.com",
                f"To={getattr(getattr(fake, 'msg', None), '__getitem__', lambda k: None)('To')} "
                f"Cc={getattr(getattr(fake, 'msg', None), '__getitem__', lambda k: None)('Cc')}",
            )
            check(
                "send_diff_email SMTP multipart/alternative HTML",
                multipart_ok,
                f"multipart={getattr(getattr(fake, 'msg', None), 'is_multipart', lambda: None)()}",
            )
        finally:
            _smtplib.SMTP = saved_smtp  # type: ignore[misc]
    finally:
        for k, v in saved_env.items():
            if v is None:
                os.environ.pop(k, None)
            else:
                os.environ[k] = v

    # 12. Excel download + parse

    print("\n--- SharePoint Excel 下載測試 ---")

    try:

        import yaml



        example_cfg = yaml.safe_load((SCRIPT_DIR / "config.example.yaml").read_text(encoding="utf-8"))

        url = example_cfg["sharepoint"]["download_url"]

        sheet = example_cfg["sharepoint"]["sheet_name"]

        cache = SCRIPT_DIR / ".register_cache"

        xlsx, original_name = sr.download_sharepoint_excel(url, cache)

        s_columns, s_rows = sr.load_register_rows_from_excel(xlsx, sheet)
        ids_s = [r.register_id for r in s_rows]
        print(f"  下載成功: {xlsx.name} ({xlsx.stat().st_size} bytes)")
        if original_name:
            print(f"  原始檔名: {original_name}")
            print(
                "  主旨括號: "
                f"[{sr.resolve_register_subject_token(original_name, example_cfg)}]"
            )
        print(f"  S 欄位: {s_columns}")
        print(f"  S 有效列數: {len(s_rows)}")

        print(f"  前 5 個 ID: {ids_s[:5]}")

        check("Excel 下載與解析", len(s_rows) > 0, f"{len(s_rows)} 列")
        check(
            "download_sharepoint_excel 回傳 (path, original_name)",
            isinstance(original_name, str),
            repr(original_name),
        )



        rf01 = next((r for r in s_rows if r.register_id == "RF-01"), None)

        rf13 = next((r for r in s_rows if r.register_id == "RF-13"), None)

        check(

            "RF-01 欄位對齊",

            rf01 is not None
            and rf01.workstream == "EVT1 RF testing"
            and rf01.title == "BT LE-2M Rx sensitivity failing"
            and rf01.next_action.startswith("QSI to investigate")
            and rf01.ball_with == "QSI"
            and rf01.priority == "P1"
            and rf01.status == "Open"
            and rf01.opened == "2026-06-22",

            f"next={rf01.next_action if rf01 else None!r}, ball={rf01.ball_with if rf01 else None!r}, "

            f"pri={rf01.priority if rf01 else None!r}, st={rf01.status if rf01 else None!r}",

        )

        check(

            "RF-13 欄位對齊",

            rf13 is not None
            and rf13.ball_with == "VOX"
            and rf13.priority == "P1"
            and rf13.status == "Closed - superseded"
            and rf13.opened == "2026-06-23",

            f"ball={rf13.ball_with if rf13 else None!r}, pri={rf13.priority if rf13 else None!r}, "

            f"st={rf13.status if rf13 else None!r}",

        )



        for r in s_rows:

            r.jira_key = f"PMWC-FAKE-{r.register_id}"

        c_html = sr.build_confluence_html_table(s_rows, site, s_columns)

        c_data_rows = c_html.count("<tr>") - 1

        check(

            "1. C 列數 = S 列數（Excel）",

            c_data_rows == len(s_rows),

            f"S={len(s_rows)}, C={c_data_rows}",

        )

    except Exception as exc:

        check("Excel 下載與解析", False, str(exc))



    print("\n" + "=" * 60)

    print(f"結果: {PASS} PASS, {FAIL} FAIL")

    print("=" * 60)

    return 0 if FAIL == 0 else 1





if __name__ == "__main__":

    sys.exit(main())


