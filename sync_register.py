#!/usr/bin/env python3
"""
SharePoint Register (S) -> Confluence mail_checking (C) -> Jira PMWC sync.

S 表格為唯一資料來源；C 與 Jira 完全跟隨 S。
"""

from __future__ import annotations

import argparse
import html
import json
import os
import re
import smtplib
import sys
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from email.message import EmailMessage
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlparse

import requests
import yaml
from dotenv import load_dotenv
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Config & constants
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
# 允許 EVT2-01 這類「字母+數字」前綴（舊版僅允許純字母，導致 EVT2 項目被忽略）
REGISTER_ID_RE = re.compile(r"^[A-Z][A-Z0-9]*-\d+$", re.IGNORECASE)
JIRA_KEY_RE = re.compile(r"PMWC-\d+")
CONFLUENCE_JIRA_LINK_RE = re.compile(
    r"\[?(PMWC-\d+)\]?\(?https?://[^)\s|]+/browse/(PMWC-\d+)\)?"
)
# S 表重複 Register ID 時，Jira summary 最前方標記（格式：ID重複 {ID}_{Title}）
DUPLICATE_ID_SUMMARY_PREFIX = "ID重複"

S_COLUMNS = [
    "ID",
    "Workstream",
    "Title",
    "Description / current state",
    "Next action",
    "Ball with",
    "Priority",
    "Status",
    "Opened",
    "Target close",
    "Next milestone / due",
    "Source / reference",
]
LINK_COLUMN = "LINK"
C_COLUMNS = S_COLUMNS + [LINK_COLUMN]

# 表頭名稱 -> RegisterRow 屬性（可選欄位如 JIRA / QSI Comment 不在現行 S 表時留空）
S_COLUMN_ATTR: dict[str, str] = {
    "ID": "register_id",
    "JIRA": "jira",
    "Workstream": "workstream",
    "Title": "title",
    "Description / current state": "description",
    "Next action": "next_action",
    "Ball with": "ball_with",
    "Priority": "priority",
    "Status": "status",
    "Opened": "opened",
    "Target close": "target_close",
    "Next milestone / due": "next_milestone",
    "Source / reference": "source",
    "QSI Comment": "qsi_comment",
}

# S 表 Status 含下列關鍵字（大小寫不敏感）→ Jira「完成」；其餘一律不改 Jira 狀態
DONE_STATUS_KEYWORDS: tuple[str, ...] = ("done", "closed", "completed")
JIRA_DONE_STATUS = "完成"
DEFAULT_JIRA_STATUS = "待辦事項"
DEFAULT_SOURCE_LINK_TITLE = "C27 Open Issues Register (S 表格)"

# Team-managed Jira：畫面類型名與 JQL 可用名不一致（例：顯示「任務」但 JQL 需 Task；
# 「大型工作」反而要用中文）。搜尋時依序嘗試。
JIRA_ISSUETYPE_JQL_ALIASES: dict[str, str] = {
    "任務": "Task",
    "大型工作": "Epic",
    "史詩": "Epic",
    "故事": "Story",
    "錯誤": "Bug",
    "子任務": "Sub-task",
}


def _jql_quote(value: str) -> str:
    """Quote a JQL string literal; escape reserved wildcards."""
    escaped = (
        (value or "")
        .replace("\\", "\\\\")
        .replace('"', '\\"')
        .replace("?", "\\u003f")
        .replace("*", "\\u002a")
    )
    return f'"{escaped}"'


def jql_issuetype_equals(type_name: str) -> str:
    """產生單一 issuetype JQL 條件。"""
    name = (type_name or "").strip()
    if not name:
        raise ValueError("issuetype name is empty")
    # ASCII simple tokens only — never leave ?/* unquoted (JQL wildcards).
    if (
        name.isascii()
        and " " not in name
        and "-" not in name
        and "?" not in name
        and "*" not in name
    ):
        return f"issuetype = {name}"
    return f"issuetype = {_jql_quote(name)}"



def resolve_issuetype_for_api(type_name: str, *, role: str) -> str:
    """Normalize config issue type for Jira create/search.

    PMWC issue types are Chinese (任務 / 大型工作). If CONFIG_YAML encoding
    damage turns them into ???, restore the project defaults by role.
    Do NOT remap valid Chinese names to English — create API rejects Task/Epic.
    """
    name = (type_name or "").strip()
    if not name or "?" in name:
        return "任務" if role == "task" else "大型工作"
    return name


def jql_issuetype_name_candidates(type_name: str) -> list[str]:
    """回傳 JQL 要嘗試的類型名稱清單（設定值 + 別名 + 英文 fallback）。"""
    name = (type_name or "").strip()
    candidates: list[str] = []
    # Skip clearly corrupted names (e.g. secret encoding lost Chinese → ???)
    if name and set(name) != {"?"}:
        candidates.append(name)
    alias = JIRA_ISSUETYPE_JQL_ALIASES.get(name)
    if alias and alias not in candidates:
        candidates.append(alias)
    # CI/config mojibake fallback: still try common English types.
    if not candidates or "?" in name:
        for zh in ("任務", "大型工作", "Task", "Epic"):
            if zh not in candidates:
                candidates.append(zh)
    return candidates



@dataclass
class RegisterRow:
    """S 表格一列；欄位與 S 表表頭同名，LINK 僅在輸出 C 表時由 jira_key 填入。"""

    register_id: str
    jira: str = ""
    workstream: str = ""
    title: str = ""
    description: str = ""
    next_action: str = ""
    ball_with: str = ""
    priority: str = ""
    status: str = ""
    opened: str = ""
    target_close: str = ""
    next_milestone: str = ""
    source: str = ""
    qsi_comment: str = ""
    jira_key: str = ""


@dataclass
class SyncReport:
    created_epics: list[str] = field(default_factory=list)
    created_issues: list[str] = field(default_factory=list)
    updated_issues: list[str] = field(default_factory=list)
    deleted_issues: list[str] = field(default_factory=list)
    web_links_added: list[str] = field(default_factory=list)
    ranked_issues: list[str] = field(default_factory=list)
    assigned_issues: list[str] = field(default_factory=list)
    confluence_updated: bool = False
    errors: list[str] = field(default_factory=list)
    diff_log_path: Path | None = None


SNAPSHOT_FIELDS: list[tuple[str, str]] = [
    ("jira", "JIRA"),
    ("workstream", "Workstream"),
    ("title", "Title"),
    ("description", "Description / current state"),
    ("next_action", "Next action"),
    ("ball_with", "Ball with"),
    ("priority", "Priority"),
    ("status", "Status"),
    ("opened", "Opened"),
    ("target_close", "Target close"),
    ("next_milestone", "Next milestone / due"),
    ("source", "Source / reference"),
    ("qsi_comment", "QSI Comment"),
]


@dataclass
class FieldChange:
    register_id: str
    field_label: str
    old_value: str
    new_value: str


@dataclass
class RegisterDiff:
    added_ids: list[str] = field(default_factory=list)
    removed_ids: list[str] = field(default_factory=list)
    status_changes: list[FieldChange] = field(default_factory=list)
    field_changes: list[FieldChange] = field(default_factory=list)
    is_first_run: bool = False
    snapshot_saved_at: str = ""
    # register_id -> 顯示用標題（新增／移除列用）
    id_titles: dict[str, str] = field(default_factory=dict)


def load_config(path: Path) -> dict[str, Any]:
    load_dotenv(SCRIPT_DIR / ".env")
    with path.open(encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    token = __import__("os").environ.get("ATLASSIAN_API_TOKEN", "").strip()
    if not token:
        raise SystemExit("缺少 ATLASSIAN_API_TOKEN，請在 .env 設定。")
    cfg["_api_token"] = token
    return cfg


def map_status_to_jira(status: str) -> str | None:
    """將 S 表 Status 對應到 Jira 狀態；僅 Done/Closed/Completed →「完成」。

    回傳 None 表示**不要變更** Jira 狀態（Open / In Progress / Blocked / 空值等）。
    比對大小寫不敏感，採子字串包含（例：「Closed - superseded」→ 完成）。
    """
    text = (status or "").strip()
    if not text:
        return None
    lowered = text.casefold()
    for keyword in DONE_STATUS_KEYWORDS:
        if keyword in lowered:
            return JIRA_DONE_STATUS
    return None


def normalize_jira_status_name(status: str) -> str:
    """正規化 Jira transition 目標名稱（已是 Jira 狀態名時原樣比對）。"""
    return (status or "").strip()


def _excel_serial_to_datetime(value: float) -> datetime:
    return datetime(1899, 12, 30) + timedelta(days=float(value))


def parse_flexible_date(value: Any, *, default_year: int | None = None) -> str | None:
    """將各種日期格式轉為 YYYY-MM-DD；無法解析則回傳 None。

    支援：ISO、斜線/橫線、中文年月日、英文月份、Excel 序號、帶時間字串等。
    無年份的格式（如 7月17日、7/17）以 default_year 或當年度補齊。
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        return _excel_serial_to_datetime(float(value)).strftime("%Y-%m-%d")

    text = str(value).strip().replace("\u3000", " ")
    if not text:
        return None

    year_default = default_year if default_year is not None else datetime.now().year

    # Excel serial（純數字字串）；短整數（如 "7"）不當序號，避免誤判
    if re.fullmatch(r"\d+(\.\d+)?", text):
        num = float(text)
        if "." in text or num >= 1000:
            try:
                return _excel_serial_to_datetime(num).strftime("%Y-%m-%d")
            except (OverflowError, ValueError):
                return None

    # 僅對 ISO／數字年月日後綴的時間裁切，避免弄壞「7 月 31 日」「Jul 31, 2026」
    m_iso_time = re.fullmatch(
        r"(\d{4}[-/.]\d{1,2}[-/.]\d{1,2})(?:[ T]\d{1,2}:\d{2}(?::\d{2})?(?:\.\d+)?Z?)?",
        text,
    )
    text_date = m_iso_time.group(1) if m_iso_time else text
    if not text_date:
        return None

    # 2026年7月17日 / 2026 年 07 月 17 号
    m = re.fullmatch(
        r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*[日号]?",
        text_date,
    )
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).strftime(
                "%Y-%m-%d"
            )
        except ValueError:
            return None

    # 7月17日 / 7 月 31 日 / 07月17号（無年份 → 當年度）
    m = re.fullmatch(r"(\d{1,2})\s*月\s*(\d{1,2})\s*[日号]?", text_date)
    if m:
        try:
            return datetime(year_default, int(m.group(1)), int(m.group(2))).strftime(
                "%Y-%m-%d"
            )
        except ValueError:
            return None

    for fmt in (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%m/%d",
        "%m-%d-%Y",
        "%m-%d-%y",
        "%m-%d",
        "%d-%b-%Y",
        "%b %d, %Y",
        "%b %d %Y",
    ):
        try:
            parsed = datetime.strptime(text_date, fmt)
            if fmt in ("%m/%d", "%m-%d"):
                parsed = parsed.replace(year=year_default)
            return parsed.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def excel_serial_to_date(value: Any) -> str:
    return parse_flexible_date(value) or (str(value).strip() if value not in (None, "") else "")


def normalize_register_id(raw: str) -> str:
    return strip_markdown_cell(raw).strip("_").upper()


def strip_markdown_cell(text: str) -> str:
    """移除 Confluence/Markdown 的粗斜體標記。"""
    value = (text or "").strip()
    value = re.sub(r"^\*+|\*+$", "", value)
    value = re.sub(r"^_+|_+$", "", value)
    return value.strip()


def is_valid_jira_key(key: str) -> bool:
    return bool(JIRA_KEY_RE.fullmatch((key or "").strip()))


def is_blank_table_row(cells: list[str]) -> bool:
    return not any(strip_markdown_cell(cell) for cell in cells)


def escape_html(text: str) -> str:
    return html.escape((text or "").replace("\n", " "), quote=True)


def normalize_cell_text(value: Any) -> str:
    """Excel 儲存格轉顯示用文字；全形空白視為空。"""
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if not text or text.replace("\u3000", "").strip() == "":
        return ""
    return text


def s_row_to_register(row: dict[str, Any]) -> RegisterRow | None:
    rid = normalize_register_id(str(row.get("ID", "")))
    if not rid or not REGISTER_ID_RE.match(rid):
        return None

    return RegisterRow(
        register_id=rid,
        jira=normalize_cell_text(row.get("JIRA")),
        workstream=normalize_cell_text(row.get("Workstream")),
        title=normalize_cell_text(row.get("Title")),
        description=normalize_cell_text(row.get("Description / current state")),
        next_action=normalize_cell_text(row.get("Next action")),
        ball_with=normalize_cell_text(row.get("Ball with")),
        priority=normalize_cell_text(row.get("Priority")),
        status=normalize_cell_text(row.get("Status")),
        opened=excel_serial_to_date(row.get("Opened")),
        target_close=excel_serial_to_date(row.get("Target close")),
        next_milestone=normalize_cell_text(row.get("Next milestone / due")),
        source=normalize_cell_text(row.get("Source / reference")),
        qsi_comment=normalize_cell_text(row.get("QSI Comment")),
    )


def jira_epic_priority(row: RegisterRow) -> str:
    """S 表 Priority 欄 → P1/P2/P3 Epic 標籤。"""
    return row.priority or "P1"


def jira_status_source(row: RegisterRow) -> str:
    """S 表 Status 欄 → Open/Closed 等狀態字串。"""
    return row.status


# ---------------------------------------------------------------------------
# Snapshot & diff
# ---------------------------------------------------------------------------


def _normalize_snapshot_text(value: str) -> str:
    """統一換行，避免 \\r\\n / \\n 造成假差異。"""
    return (value or "").replace("\r\n", "\n").replace("\r", "\n")


def register_row_to_snapshot_dict(row: RegisterRow) -> dict[str, str]:
    return {
        attr: _normalize_snapshot_text(str(getattr(row, attr) or ""))
        for attr, _ in SNAPSHOT_FIELDS
    }


def load_last_snapshot(path: Path) -> tuple[dict[str, dict[str, str]], str]:
    if not path.exists():
        return {}, ""
    data = json.loads(path.read_text(encoding="utf-8"))
    rows_raw = data.get("rows", {})
    rows: dict[str, dict[str, str]] = {}
    for rid, row in rows_raw.items():
        if not isinstance(row, dict):
            continue
        rows[str(rid)] = {
            str(k): _normalize_snapshot_text(str(v if v is not None else ""))
            for k, v in row.items()
        }
    saved_at = str(data.get("saved_at", ""))
    return rows, saved_at


def save_last_snapshot(path: Path, rows: list[RegisterRow]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "rows": {r.register_id: register_row_to_snapshot_dict(r) for r in rows},
    }
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _snapshot_title(row: dict[str, str] | None) -> str:
    if not row:
        return ""
    return (row.get("title") or "").strip()


def compute_register_diff(
    old_rows: dict[str, dict[str, str]],
    new_rows: list[RegisterRow],
    *,
    snapshot_saved_at: str = "",
) -> RegisterDiff:
    diff = RegisterDiff()
    diff.is_first_run = not old_rows
    diff.snapshot_saved_at = snapshot_saved_at

    old_ids = set(old_rows.keys())
    new_ids = {r.register_id for r in new_rows}
    new_map = {r.register_id: register_row_to_snapshot_dict(r) for r in new_rows}

    # 標題對照（新增／移除顯示用）
    for rid, row in old_rows.items():
        title = _snapshot_title(row)
        if title:
            diff.id_titles[rid] = title
    for rid, row in new_map.items():
        title = _snapshot_title(row)
        if title:
            diff.id_titles[rid] = title

    if diff.is_first_run:
        diff.added_ids = sorted(new_ids)
        return diff

    diff.added_ids = sorted(new_ids - old_ids)
    diff.removed_ids = sorted(old_ids - new_ids)

    for rid in sorted(new_ids & old_ids):
        old_row = old_rows[rid]
        new_row = new_map[rid]
        for attr, label in SNAPSHOT_FIELDS:
            old_val = _normalize_snapshot_text(old_row.get(attr, ""))
            new_val = _normalize_snapshot_text(new_row.get(attr, ""))
            if old_val == new_val:
                continue
            change = FieldChange(rid, label, old_val, new_val)
            if attr == "status":
                diff.status_changes.append(change)
            else:
                diff.field_changes.append(change)

    return diff


def _truncate_diff_value(text: str, max_len: int = 80) -> str:
    value = (text or "").replace("\n", " ")
    if len(value) <= max_len:
        return value
    return value[: max_len - 3] + "..."


def _format_diff_value_pair(
    old: str, new: str, *, max_len: int = 80
) -> tuple[str, str]:
    """截斷顯示；若前綴相同則把視窗移到第一個差異處，避免郵件看起來「沒變」。"""
    old_s = (old or "").replace("\n", " ")
    new_s = (new or "").replace("\n", " ")
    if len(old_s) <= max_len and len(new_s) <= max_len:
        return old_s, new_s

    # 找第一個差異索引
    limit = min(len(old_s), len(new_s))
    idx = 0
    while idx < limit and old_s[idx] == new_s[idx]:
        idx += 1
    if idx == limit and len(old_s) == len(new_s):
        # 理論上不應進 diff；仍安全截斷
        return _truncate_diff_value(old_s, max_len), _truncate_diff_value(new_s, max_len)

    # 在差異點前後留 context
    context = max(12, max_len // 4)
    start = max(0, idx - context)

    def _window(s: str) -> str:
        chunk = s[start:]
        prefix = "..." if start > 0 else ""
        body_budget = max_len - len(prefix) - 3
        if len(chunk) <= body_budget + 3:
            return prefix + chunk
        return prefix + chunk[:body_budget] + "..."

    return _window(old_s), _window(new_s)


def _format_id_with_title(rid: str, titles: dict[str, str]) -> str:
    title = (titles.get(rid) or "").strip()
    if not title:
        return rid
    short = _truncate_diff_value(title, 60)
    return f"{rid} — {short}"


def format_diff_report(diff: RegisterDiff, *, dry_run: bool = False) -> str:
    lines: list[str] = ["========== S 表格變更摘要 =========="]
    titles = diff.id_titles or {}

    if diff.is_first_run:
        lines.append("比對基準: 首次執行（無前次快照）")
    else:
        saved = diff.snapshot_saved_at or "未知"
        lines.append(f"比對基準: 上次同步快照 ({saved})")

    if dry_run:
        lines.append("模式: dry-run（預覽，不寫入）")

    lines.append("")

    if diff.is_first_run:
        lines.append(f"【首次快照】目前 S 表格共 {len(diff.added_ids)} 筆有效 ID")
        for rid in diff.added_ids:
            lines.append(f"  • {_format_id_with_title(rid, titles)}")
        lines.append("")
        lines.append("【結論】首次執行，僅建立基準快照（尚無前後差異可比）。")
        return "\n".join(lines)

    lines.append(f"【新增】{len(diff.added_ids)} 筆")
    if diff.added_ids:
        for rid in diff.added_ids:
            lines.append(f"  + {_format_id_with_title(rid, titles)}")
    else:
        lines.append("  （無）")

    lines.append("")
    lines.append(f"【移除】{len(diff.removed_ids)} 筆")
    if diff.removed_ids:
        for rid in diff.removed_ids:
            lines.append(f"  - {_format_id_with_title(rid, titles)}")
    else:
        lines.append("  （無）")

    lines.append("")
    lines.append(f"【狀態變更】{len(diff.status_changes)} 筆")
    if diff.status_changes:
        for change in diff.status_changes:
            old = change.old_value or "（空）"
            new = change.new_value or "（空）"
            label = _format_id_with_title(change.register_id, titles)
            lines.append(f"  {label}: {old} → {new}")
    else:
        lines.append("  （無）")

    lines.append("")
    lines.append(f"【欄位變更】{len(diff.field_changes)} 筆")
    if diff.field_changes:
        for change in diff.field_changes:
            old, new = _format_diff_value_pair(change.old_value, change.new_value)
            old_disp = old or "（空）"
            new_disp = new or "（空）"
            lines.append(
                f"  {change.register_id} / {change.field_label}: "
                f"{old_disp!r} → {new_disp!r}"
            )
    else:
        lines.append("  （無）")

    total = (
        len(diff.added_ids)
        + len(diff.removed_ids)
        + len(diff.status_changes)
        + len(diff.field_changes)
    )
    lines.append("")
    if total == 0:
        lines.append("【結論】無差異")
        lines.append("本次 S 表格與上次快照完全相同，無任何變更（含狀態）。")
    else:
        lines.append(f"【結論】有差異（共 {total} 項變更）")

    return "\n".join(lines)


def diff_has_changes(diff: RegisterDiff) -> bool:
    if diff.is_first_run:
        return False
    return bool(
        diff.added_ids
        or diff.removed_ids
        or diff.status_changes
        or diff.field_changes
    )


def resolve_diff_email_verdict(diff: RegisterDiff) -> str:
    """郵件主旨／內文第一行結論：首次快照 | 有差異 | 無差異。"""
    if diff.is_first_run:
        return "首次快照"
    if diff_has_changes(diff):
        return "有差異"
    return "無差異"


def write_sync_log(
    log_dir: Path,
    diff: RegisterDiff,
    diff_text: str,
    report: SyncReport,
) -> Path:
    log_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = log_dir / f"sync_{timestamp}.log"
    content = "\n\n".join([diff_text, _format_sync_result_section(report)])
    log_path.write_text(content + "\n", encoding="utf-8")

    # 固定檔名，方便本機／自動化取用
    latest_path = log_dir / "latest_diff.txt"
    latest_path.write_text(content + "\n", encoding="utf-8")
    root_latest = SCRIPT_DIR / "最新同步差異.txt"
    root_latest.write_text(content + "\n", encoding="utf-8")

    summary_path = log_dir / "sync_log.txt"
    summary_line = (
        f"{datetime.now().isoformat(timespec='seconds')} | {log_path.name} | "
        f"新增={len(diff.added_ids)} 移除={len(diff.removed_ids)} "
        f"狀態={len(diff.status_changes)} 欄位={len(diff.field_changes)} "
        f"錯誤={len(report.errors)}\n"
    )
    with summary_path.open("a", encoding="utf-8") as f:
        f.write(summary_line)

    return log_path


def is_github_actions() -> bool:
    return os.environ.get("GITHUB_ACTIONS", "").lower() == "true"


def build_diff_email_body(
    diff_text: str,
    report: SyncReport,
    *,
    verdict: str = "",
    s_link_url: str = "",
    s_link_title: str = "",
    confluence_url: str = "",
) -> str:
    """純文字郵件內文。第一行為【有差異】／【無差異】／【首次快照】。"""
    lines: list[str] = []
    if verdict:
        lines.append(f"【{verdict}】")
        lines.append("")
    lines.extend(
        [
            "PMWC Register 同步差異報告",
            f"時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"執行環境: {'GitHub Actions' if is_github_actions() else '本機'}",
            "",
        ]
    )
    if s_link_url:
        lines.append(f"S 表格連結: {s_link_title or 'SharePoint Register'}")
        lines.append(s_link_url)
        lines.append("")
    if confluence_url:
        lines.append(f"C 表格 (Confluence): {confluence_url}")
        lines.append("")
    lines.append(diff_text)
    lines.append("")
    lines.append(_format_sync_result_section(report))
    return "\n".join(lines)


def build_diff_email_html_body(plain_body: str, *, verdict: str) -> str:
    """HTML 內文：有差異時第一行紅色粗體；其餘以 escaped <pre> 呈現。"""
    if verdict == "有差異":
        banner = (
            '<div style="font-size:1.25em;margin:0 0 12px 0">'
            '<span style="color:red;font-weight:bold">有差異</span>'
            "</div>"
        )
    elif verdict:
        banner = (
            '<div style="font-size:1.25em;font-weight:bold;margin:0 0 12px 0">'
            f"{html.escape(verdict)}"
            "</div>"
        )
    else:
        banner = ""

    rest = plain_body
    if verdict:
        # 略過純文字第一行 【verdict】與緊接空行，避免 HTML 重複標題
        lines = plain_body.splitlines()
        if lines and lines[0].strip() in {f"【{verdict}】", verdict}:
            lines = lines[1:]
            if lines and lines[0].strip() == "":
                lines = lines[1:]
        rest = "\n".join(lines)

    escaped = html.escape(rest)
    return (
        "<!DOCTYPE html><html><body "
        'style="font-family:Segoe UI,Arial,sans-serif;font-size:14px">'
        f"{banner}"
        '<pre style="white-space:pre-wrap;font-family:Consolas,monospace;'
        f'font-size:13px;margin:0">{escaped}</pre>'
        "</body></html>"
    )


def _email_domain(addr: str) -> str:
    addr = (addr or "").strip().lower()
    if "@" not in addr:
        return ""
    return addr.rsplit("@", 1)[-1].strip()


def normalize_email_addrs(value: Any) -> list[str]:
    """將 email_to / email_cc（字串、list、逗號／分號分隔）正規成地址清單。"""
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        raw_items = [str(x) for x in value]
    else:
        raw_items = re.split(r"[,;\s]+", str(value))
    out: list[str] = []
    seen: set[str] = set()
    for item in raw_items:
        addr = item.strip()
        if not addr or "@" not in addr:
            continue
        key = addr.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(addr)
    return out


def resolve_notify_recipients(cfg: dict[str, Any]) -> tuple[list[str], list[str]]:
    """回傳 (To, Cc)。To／Cc 分離；勿把 Cc 併入 To。

    To：SYNC_NOTIFY_EMAIL → notify.email_to → bob.lin@qsitw.com
    Cc：SYNC_NOTIFY_CC → notify.email_cc（可為字串或 list）
    """
    notify = cfg.get("notify") or {}
    to_raw = os.environ.get("SYNC_NOTIFY_EMAIL")
    if to_raw is None or not str(to_raw).strip():
        to_raw = notify.get("email_to")
    to_addrs = normalize_email_addrs(to_raw)
    if not to_addrs:
        to_addrs = ["bob.lin@qsitw.com"]

    cc_raw = os.environ.get("SYNC_NOTIFY_CC")
    if cc_raw is None or not str(cc_raw).strip():
        cc_raw = notify.get("email_cc")
    cc_addrs = normalize_email_addrs(cc_raw)
    # 避免同一人同時出現在 To 與 Cc
    to_keys = {a.lower() for a in to_addrs}
    cc_addrs = [a for a in cc_addrs if a.lower() not in to_keys]
    return to_addrs, cc_addrs


def infer_smtp_host(username: str) -> str:
    """依寄件帳號網域推斷 SMTP（my-bot 其他腳本對 Gmail 用 smtp.gmail.com）。"""
    domain = _email_domain(username)
    if domain in {"gmail.com", "googlemail.com"}:
        return "smtp.gmail.com"
    if domain in {"outlook.com", "hotmail.com", "live.com", "msn.com"}:
        return "smtp.office365.com"
    # 企業網域預設走 Microsoft 365（若 Basic Auth 停用需改 Graph）
    return "smtp.office365.com"


def graph_credentials_present() -> bool:
    tenant = (
        os.environ.get("GRAPH_TENANT_ID")
        or os.environ.get("AZURE_TENANT_ID")
        or ""
    ).strip()
    client_id = (
        os.environ.get("GRAPH_CLIENT_ID")
        or os.environ.get("AZURE_CLIENT_ID")
        or ""
    ).strip()
    client_secret = (
        os.environ.get("GRAPH_CLIENT_SECRET")
        or os.environ.get("AZURE_CLIENT_SECRET")
        or ""
    ).strip()
    return bool(tenant and client_id and client_secret)


def resolve_mail_backend(cfg: dict[str, Any]) -> str:
    """回傳 mail backend：graph | smtp。

    notify.mail_backend / MAIL_BACKEND：auto|smtp|graph
    auto：有 GRAPH_/AZURE_ client credentials 則用 graph，否則 smtp。
    """
    notify = cfg.get("notify") or {}
    raw = (
        os.environ.get("MAIL_BACKEND")
        or notify.get("mail_backend")
        or "auto"
    )
    raw = str(raw).strip().lower() or "auto"
    if raw == "auto":
        return "graph" if graph_credentials_present() else "smtp"
    if raw in {"smtp", "graph"}:
        return raw
    raise RuntimeError(
        f"未知 mail_backend={raw!r}（允許 auto|smtp|graph）"
    )


def build_graph_send_mail_payload(
    *,
    subject: str,
    body: str,
    to_addr: str | list[str],
    save_to_sent: bool = False,
    html_body: str | None = None,
    cc_addrs: str | list[str] | None = None,
) -> dict[str, Any]:
    """組 Microsoft Graph users/.../sendMail JSON（供單元測試 mock）。

    有 html_body 時以 HTML 寄出（紅色「有差異」等）；否則 Text。
    Cc 走 ccRecipients，不可併入 toRecipients。
    """
    if html_body:
        content_type = "HTML"
        content = html_body
    else:
        content_type = "Text"
        content = body
    to_list = normalize_email_addrs(to_addr)
    if not to_list:
        raise ValueError("build_graph_send_mail_payload: To 收件人不可為空")
    message: dict[str, Any] = {
        "subject": subject,
        "body": {"contentType": content_type, "content": content},
        "toRecipients": [
            {"emailAddress": {"address": addr}} for addr in to_list
        ],
    }
    cc_list = normalize_email_addrs(cc_addrs)
    if cc_list:
        message["ccRecipients"] = [
            {"emailAddress": {"address": addr}} for addr in cc_list
        ]
    return {
        "message": message,
        "saveToSentItems": save_to_sent,
    }


def _graph_token_endpoint(tenant: str) -> str:
    return f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"


def acquire_graph_token(
    *,
    tenant: str,
    client_id: str,
    client_secret: str,
    timeout: float = 60,
) -> str:
    """Client credentials 取得 Graph access token。"""
    resp = requests.post(
        _graph_token_endpoint(tenant),
        data={
            "grant_type": "client_credentials",
            "client_id": client_id,
            "client_secret": client_secret,
            "scope": "https://graph.microsoft.com/.default",
        },
        timeout=timeout,
    )
    if resp.status_code >= 400:
        raise RuntimeError(
            f"Graph token 取得失敗 HTTP {resp.status_code}: {resp.text[:500]}"
        )
    token = (resp.json() or {}).get("access_token") or ""
    if not token:
        raise RuntimeError("Graph token 回應缺少 access_token")
    return token


def send_diff_email_graph(
    cfg: dict[str, Any],
    *,
    subject: str,
    body: str,
    html_body: str | None = None,
) -> None:
    """以 Microsoft Graph sendMail 寄信（需 Mail.Send application permission）。"""
    notify = cfg.get("notify") or {}
    to_addrs, cc_addrs = resolve_notify_recipients(cfg)
    to_addr = ", ".join(to_addrs)
    cc_disp = ", ".join(cc_addrs) if cc_addrs else "(none)"
    from_addr = (
        os.environ.get("GRAPH_MAILBOX")
        or os.environ.get("MAIL_USERNAME")
        or notify.get("smtp_user")
        or notify.get("graph_mailbox")
        or ""
    ).strip()
    tenant = (
        os.environ.get("GRAPH_TENANT_ID")
        or os.environ.get("AZURE_TENANT_ID")
        or notify.get("graph_tenant_id")
        or ""
    ).strip()
    client_id = (
        os.environ.get("GRAPH_CLIENT_ID")
        or os.environ.get("AZURE_CLIENT_ID")
        or notify.get("graph_client_id")
        or ""
    ).strip()
    client_secret = (
        os.environ.get("GRAPH_CLIENT_SECRET")
        or os.environ.get("AZURE_CLIENT_SECRET")
        or notify.get("graph_client_secret")
        or ""
    ).strip()

    if not from_addr:
        raise RuntimeError(
            "Graph 寄信失敗：缺少 GRAPH_MAILBOX / MAIL_USERNAME（寄件信箱）"
        )
    if not (tenant and client_id and client_secret):
        raise RuntimeError(
            "Graph 寄信失敗：缺少 GRAPH_TENANT_ID / GRAPH_CLIENT_ID / "
            "GRAPH_CLIENT_SECRET（或 AZURE_* 同義）"
        )

    print(
        f"Graph 準備寄信: mailbox={from_addr} to={to_addr} cc={cc_disp} "
        f"tenant_set=True client_id_set=True"
    )
    token = acquire_graph_token(
        tenant=tenant, client_id=client_id, client_secret=client_secret
    )
    payload = build_graph_send_mail_payload(
        subject=subject,
        body=body,
        to_addr=to_addrs,
        html_body=html_body,
        cc_addrs=cc_addrs,
    )
    url = (
        "https://graph.microsoft.com/v1.0/users/"
        f"{requests.utils.quote(from_addr, safe='@')}/sendMail"
    )
    resp = requests.post(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=60,
    )
    if resp.status_code not in (202, 200):
        raise RuntimeError(
            f"Graph sendMail 失敗 HTTP {resp.status_code}: {resp.text[:800]}"
        )


def send_diff_email_smtp(
    cfg: dict[str, Any],
    *,
    subject: str,
    body: str,
    html_body: str | None = None,
) -> None:
    """透過 SMTP 寄出差異報告（GitHub Secrets: MAIL_USERNAME / MAIL_PASSWORD）。

    主機推斷：明確 SMTP_HOST / notify.smtp_host 優先；否則依 MAIL_USERNAME 網域
    （gmail → smtp.gmail.com；其餘預設 smtp.office365.com）。
    my-bot 內 guardian_bot / DailyStockPush 使用同一組 MAIL_* + smtp.gmail.com。
    From 固定為 MAIL_USERNAME。
    有 html_body 時以 multipart/alternative（text + html）寄出。
    """
    notify = cfg.get("notify") or {}
    to_addrs, cc_addrs = resolve_notify_recipients(cfg)
    to_addr = ", ".join(to_addrs)
    cc_disp = ", ".join(cc_addrs) if cc_addrs else "(none)"
    user = (
        os.environ.get("MAIL_USERNAME")
        or notify.get("smtp_user")
        or ""
    ).strip()
    password = (
        os.environ.get("MAIL_PASSWORD")
        or notify.get("smtp_password")
        or ""
    ).strip()
    host = (
        os.environ.get("SMTP_HOST")
        or notify.get("smtp_host")
        or (infer_smtp_host(user) if user else "")
        or "smtp.gmail.com"
    ).strip()
    port = int(os.environ.get("SMTP_PORT") or notify.get("smtp_port") or 587)
    from_addr = user

    if not user or not password:
        raise RuntimeError(
            "寄信失敗：缺少 MAIL_USERNAME / MAIL_PASSWORD（或 notify.smtp_*）"
        )

    user_domain = _email_domain(user) or "?"
    print(
        f"SMTP 準備寄信: host={host}:{port} user_domain={user_domain} "
        f"from={from_addr} to={to_addr} cc={cc_disp} password_set=True "
        f"password_len={len(password)}"
    )

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = from_addr
    msg["To"] = to_addr
    if cc_addrs:
        msg["Cc"] = ", ".join(cc_addrs)
    msg.set_content(body)
    if html_body:
        msg.add_alternative(html_body, subtype="html")

    try:
        with smtplib.SMTP(host, port, timeout=60) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.ehlo()
            smtp.login(user, password)
            smtp.send_message(msg)
    except smtplib.SMTPAuthenticationError as exc:
        err = str(exc)
        hint = ""
        if "5.7.139" in err or "basic authentication is disabled" in err.lower():
            hint = (
                " 提示：目標伺服器已停用 SMTP Basic Auth。"
                "若 MAIL_USERNAME 為 Gmail，請設 SMTP_HOST=smtp.gmail.com"
                "（或依賴網域自動推斷）；"
                "若為 Microsoft 365，請改 Graph"
                "（GRAPH_TENANT_ID/CLIENT_ID/CLIENT_SECRET + Mail.Send）"
                "或啟用 Authenticated SMTP / 應用程式密碼。"
            )
        elif "gmail" in host.lower():
            hint = (
                " 提示：請確認 MAIL_USERNAME / MAIL_PASSWORD 為"
                " Gmail 應用程式密碼（與 guardian_bot 相同）。"
            )
        raise RuntimeError(
            f"SMTP 認證失敗 ({host}:{port}, user_domain={user_domain}): "
            f"{exc}.{hint}"
        ) from exc
    except smtplib.SMTPException as exc:
        raise RuntimeError(
            f"SMTP 寄信失敗 ({host}:{port}, user_domain={user_domain}, "
            f"from={from_addr}, to={to_addr}): {exc}"
        ) from exc


def send_diff_email(
    cfg: dict[str, Any],
    *,
    subject: str,
    body: str,
    html_body: str | None = None,
) -> None:
    """依 notify.mail_backend（auto|smtp|graph）寄出差異報告。"""
    backend = resolve_mail_backend(cfg)
    print(f"郵件後端: {backend}")
    if backend == "graph":
        send_diff_email_graph(
            cfg, subject=subject, body=body, html_body=html_body
        )
    else:
        send_diff_email_smtp(
            cfg, subject=subject, body=body, html_body=html_body
        )


def deliver_diff_report(
    cfg: dict[str, Any],
    log_dir: Path,
    register_diff: RegisterDiff,
    report: SyncReport,
    *,
    dry_run: bool,
    s_link_url: str = "",
    s_link_title: str = "",
    confluence_url: str = "",
    s_original_filename: str = "",
    s_previous_filename: str = "",
) -> str:
    """同步後必做：寫入本機差異檔；GitHub Actions 另寄郵件（無差異也寄）。"""
    diff_text = print_diff_report(register_diff, dry_run=dry_run)
    report.diff_log_path = write_sync_log(log_dir, register_diff, diff_text, report)
    print(f"變更紀錄已寫入: {report.diff_log_path}")
    print(f"最新差異副本: {SCRIPT_DIR / '最新同步差異.txt'}")

    notify = cfg.get("notify") or {}
    send_mail = bool(notify.get("email_on_github", True)) and is_github_actions()
    if notify.get("email_always"):
        send_mail = True
    if dry_run:
        send_mail = False

    if send_mail:
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        verdict = resolve_diff_email_verdict(register_diff)
        subject_token = resolve_register_subject_token(
            s_original_filename,
            cfg,
            previous_filename=s_previous_filename,
        )
        subject = build_diff_email_subject(
            subject_token=subject_token, verdict=verdict, stamp=stamp
        )
        body = build_diff_email_body(
            diff_text,
            report,
            verdict=verdict,
            s_link_url=s_link_url,
            s_link_title=s_link_title,
            confluence_url=confluence_url,
        )
        html_body = build_diff_email_html_body(body, verdict=verdict)
        try:
            send_diff_email(
                cfg, subject=subject, body=body, html_body=html_body
            )
            print(f"差異報告已寄出: {subject}")
        except Exception as exc:  # noqa: BLE001
            report.errors.append(f"寄送差異郵件: {exc}")
            print(f"警告: 寄送差異郵件失敗: {exc}")
            if is_github_actions():
                report.errors.append(
                    "GITHUB_ACTIONS 寄信失敗，將以非零結束碼退出"
                )

    return diff_text


def _format_sync_result_section(report: SyncReport) -> str:
    lines = ["========== 同步執行結果 =========="]
    if report.created_epics:
        lines.append(f"新建 Epic: {', '.join(report.created_epics)}")
    if report.created_issues:
        lines.append(f"新建任務: {', '.join(report.created_issues)}")
    if report.updated_issues:
        lines.append(f"更新任務: {', '.join(report.updated_issues)}")
    if report.deleted_issues:
        lines.append(f"刪除任務: {', '.join(report.deleted_issues)}")
    if report.web_links_added:
        lines.append(f"新增 Web 連結: {', '.join(report.web_links_added)}")
    if report.ranked_issues:
        lines.append(f"重排 Rank: {len(report.ranked_issues)} 筆")
    if report.assigned_issues:
        lines.append(f"指派受託人: {', '.join(report.assigned_issues)}")
    lines.append(
        f"Confluence 已更新: {'是' if report.confluence_updated else '否（dry-run 或未變更）'}"
    )
    if report.errors:
        lines.append("")
        lines.append("錯誤:")
        for err in report.errors:
            lines.append(f"  - {err}")
    return "\n".join(lines)


def print_diff_report(diff: RegisterDiff, *, dry_run: bool = False) -> str:
    text = format_diff_report(diff, dry_run=dry_run)
    print(f"\n{text}")
    return text


# ---------------------------------------------------------------------------
# SharePoint / Excel
# ---------------------------------------------------------------------------

_REGISTER_CACHE_NAME = "register_latest.xlsx"
_REGISTER_ORIGINAL_NAME_FILE = "register_original_name.txt"
_EXCEL_SUFFIXES = (".xlsx", ".xls", ".xlsm")


def parse_content_disposition_filename(header: str) -> str | None:
    """從 Content-Disposition 取出 basename（優先 filename*）。"""
    if not (header or "").strip():
        return None
    # RFC 5987：filename*=UTF-8''percent-encoded
    m_star = re.search(r"filename\*\s*=\s*([^;]+)", header, flags=re.IGNORECASE)
    if m_star:
        raw = m_star.group(1).strip().strip('"').strip("'")
        if "'" in raw:
            parts = raw.split("'", 2)
            if len(parts) == 3:
                raw = parts[2]
        name = unquote(raw).strip().strip('"').strip("'")
        if name:
            return Path(name).name
    m = re.search(r"filename\s*=\s*([^;]+)", header, flags=re.IGNORECASE)
    if m:
        raw = m.group(1).strip().strip('"').strip("'")
        if raw.lower().startswith("utf-8''"):
            raw = unquote(raw[7:])
        name = raw.strip()
        if name:
            return Path(name).name
    return None


def filename_from_download_url(url: str) -> str | None:
    """若 URL path 以 .xlsx/.xls/.xlsm 結尾則取 basename。"""
    path = unquote(urlparse(url or "").path or "")
    name = Path(path).name
    if name and name.lower().endswith(_EXCEL_SUFFIXES):
        return name
    return None


def extract_download_original_filename(
    resp: requests.Response, request_url: str
) -> str:
    """Prefer Content-Disposition，其次最終／請求 URL 的 .xlsx 檔名。"""
    header = resp.headers.get("Content-Disposition") or resp.headers.get(
        "content-disposition"
    )
    for candidate in (
        parse_content_disposition_filename(header or ""),
        filename_from_download_url(str(resp.url or "")),
        filename_from_download_url(request_url),
    ):
        if candidate and candidate.lower() != _REGISTER_CACHE_NAME:
            return candidate
    return ""


def filename_stem_for_subject(name: str) -> str:
    """去掉路徑與 Excel 副檔名，供郵件主旨 [bracket] 使用。"""
    base = Path((name or "").strip()).name
    if not base:
        return ""
    lower = base.lower()
    for ext in _EXCEL_SUFFIXES:
        if lower.endswith(ext):
            return base[: -len(ext)]
    return base


def sanitize_subject_token(token: str) -> str:
    """清理括號內字元；空白改底線。"""
    t = (token or "").strip()
    t = re.sub(r'[\r\n\[\]<>:"/\\|?*]', "", t)
    t = re.sub(r"\s+", "_", t)
    return t.strip("._")


def format_filename_transition(
    old_name: str,
    new_name: str,
    *,
    head_len: int = 3,
    min_tail: int = 4,
) -> str:
    """郵件主旨用檔名過渡字串。

    相同／缺舊名 → 僅目前 stem。
    有重新命名 → 例如 ``C27...0702->C27...0704``（共用前綴縮成 head+...，差異尾段保留）。
    """
    old_s = sanitize_subject_token(filename_stem_for_subject(old_name))
    new_s = sanitize_subject_token(filename_stem_for_subject(new_name))
    if not new_s and not old_s:
        return ""
    if not old_s or old_s == new_s:
        return new_s or old_s
    if not new_s:
        return old_s

    lcp = 0
    limit = min(len(old_s), len(new_s))
    while lcp < limit and old_s[lcp] == new_s[lcp]:
        lcp += 1

    lcs = 0
    max_lcs = min(len(old_s) - lcp, len(new_s) - lcp)
    while lcs < max_lcs and old_s[-(lcs + 1)] == new_s[-(lcs + 1)]:
        lcs += 1

    old_core = old_s[lcp : len(old_s) - lcs] if lcs else old_s[lcp:]
    new_core = new_s[lcp : len(new_s) - lcs] if lcs else new_s[lcp:]

    # 差異段太短時往左併入共用前綴（0702 而非僅 2）
    while lcp > 0 and max(len(old_core), len(new_core)) < min_tail:
        lcp -= 1
        old_core = old_s[lcp] + old_core
        new_core = new_s[lcp] + new_core

    suffix = old_s[len(old_s) - lcs :] if lcs else ""
    if lcp > 0:
        head = old_s[: min(head_len, lcp)]
        old_part = f"{head}...{old_core}{suffix}"
        new_part = f"{head}...{new_core}{suffix}"
        return f"{old_part}->{new_part}"
    return f"{old_s}->{new_s}"


def read_cached_register_original_name(cache_dir: Path) -> str:
    """讀取上次同步寫入的 S 表原始檔名（下載前呼叫以保留 previous）。"""
    path = cache_dir / _REGISTER_ORIGINAL_NAME_FILE
    if not path.is_file():
        return ""
    return path.read_text(encoding="utf-8").strip()


def resolve_register_subject_token(
    original_filename: str,
    cfg: dict[str, Any] | None = None,
    *,
    previous_filename: str = "",
) -> str:
    """郵件主旨括號 token：優先真實 S 表檔名（去 .xlsx）；有改名則縮寫舊→新。"""
    cfg = cfg or {}
    sp = cfg.get("sharepoint") or {}
    conf = cfg.get("confluence") or {}
    candidates = (
        original_filename,
        str(sp.get("register_filename") or ""),
        str(conf.get("source_link_title") or ""),
        DEFAULT_SOURCE_LINK_TITLE,
        "Open_Issues_Register",
    )
    current_token = ""
    for cand in candidates:
        stem = filename_stem_for_subject(str(cand))
        token = sanitize_subject_token(stem)
        if not token:
            continue
        if token.lower() == "register_latest":
            continue
        if "?" in token or "\ufffd" in token:
            continue
        current_token = token
        break
    if not current_token:
        current_token = "Open_Issues_Register"

    prev = (previous_filename or "").strip()
    if not prev:
        return current_token
    new_for_cmp = (original_filename or "").strip() or current_token
    transition = format_filename_transition(prev, new_for_cmp)
    return transition or current_token


def build_diff_email_subject(
    *,
    subject_token: str,
    verdict: str,
    stamp: str | None = None,
) -> str:
    """例：[C27...0702->C27...0704] 有差異 — 2026-07-15 11:00"""
    when = stamp or datetime.now().strftime("%Y-%m-%d %H:%M")
    return f"[{subject_token}] {verdict} — {when}"


def resolve_local_xlsx_override() -> Path | None:
    """環境變數 SYNC_LOCAL_XLSX：改讀本機 Excel（跳過 SharePoint 下載）。"""
    raw = (os.environ.get("SYNC_LOCAL_XLSX") or "").strip().strip('"')
    if not raw:
        return None
    path = Path(raw).expanduser()
    if not path.is_file():
        raise RuntimeError(f"SYNC_LOCAL_XLSX 不是有效檔案: {path}")
    return path.resolve()


def download_sharepoint_excel(url: str, cache_dir: Path) -> tuple[Path, str]:
    """下載至 register_latest.xlsx；回傳 (path, 原始檔名 basename)。

    原始檔名優先取自 Content-Disposition（filename / filename*），否則 URL。
    同時寫入 cache_dir/register_original_name.txt（若有解析到）。

    若設 SYNC_LOCAL_XLSX，改用本機檔並複製到 cache（便於驗證郵件／差異）。
    """
    cache_dir.mkdir(parents=True, exist_ok=True)
    dest = cache_dir / _REGISTER_CACHE_NAME
    name_file = cache_dir / _REGISTER_ORIGINAL_NAME_FILE

    local = resolve_local_xlsx_override()
    if local is not None:
        data = local.read_bytes()
        if len(data) < 1000:
            raise RuntimeError(f"SYNC_LOCAL_XLSX 檔案過小: {local}")
        dest.write_bytes(data)
        original = local.name
        name_file.write_text(original, encoding="utf-8")
        print(f"使用本機 S 表 (SYNC_LOCAL_XLSX): {local}")
        print(f"S 表原始檔名: {original}")
        return dest, original

    resp = requests.get(url, timeout=120, allow_redirects=True)
    resp.raise_for_status()
    if len(resp.content) < 1000:
        raise RuntimeError("下載的檔案過小，可能不是有效的 xlsx。")
    dest.write_bytes(resp.content)
    original = extract_download_original_filename(resp, url)
    if original:
        name_file.write_text(original, encoding="utf-8")
        print(f"S 表原始檔名: {original}")
    elif name_file.is_file():
        cached = name_file.read_text(encoding="utf-8").strip()
        if cached:
            original = cached
            print(f"S 表原始檔名（快取）: {original}")
    return dest, original


def _cell_column_letter(cell: Any) -> str | None:
    letter = getattr(cell, "column_letter", None)
    return letter if letter else None


def build_column_letter_header_map(header_cells: tuple[Any, ...]) -> dict[str, str]:
    """表頭列：欄位字母 (A, B, C…) -> 表頭名稱。"""
    mapping: dict[str, str] = {}
    for cell in header_cells:
        letter = _cell_column_letter(cell)
        if not letter or cell.value is None:
            continue
        label = str(cell.value).strip()
        if label:
            mapping[letter] = label
    return mapping


def excel_row_to_dict_by_column_letter(
    row_cells: tuple[Any, ...],
    col_letter_to_header: dict[str, str],
) -> dict[str, Any]:
    """以儲存格欄位字母對應表頭，避免 read_only 稀疏列造成 tuple 索引錯位。"""
    row_dict: dict[str, Any] = {}
    for cell in row_cells:
        letter = _cell_column_letter(cell)
        if not letter:
            continue
        header = col_letter_to_header.get(letter)
        if header:
            row_dict[header] = cell.value
    return row_dict


def load_s_column_headers(header_cells: tuple[Any, ...]) -> list[str]:
    """依 Excel 表頭列順序（A, B, C…）回傳欄位名稱。"""
    headers: list[str] = []
    for cell in header_cells:
        if cell.value is None:
            continue
        label = str(cell.value).strip()
        if label:
            headers.append(label)
    return headers


def c_columns_for(s_columns: list[str]) -> list[str]:
    return s_columns + [LINK_COLUMN]


def cell_value_for_s_column(row: RegisterRow, column: str) -> str:
    if column == "ID":
        return row.register_id
    attr = S_COLUMN_ATTR.get(column)
    if not attr:
        return ""
    return str(getattr(row, attr) or "")


def load_register_rows_from_excel(
    xlsx_path: Path, sheet_name: str
) -> tuple[list[str], list[RegisterRow]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"找不到工作表 {sheet_name!r}，現有：{wb.sheetnames}")
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows()
    header_cells = next(rows_iter, None)
    if not header_cells:
        wb.close()
        return [], []
    s_columns = load_s_column_headers(header_cells)
    col_letter_to_header = build_column_letter_header_map(header_cells)
    result: list[RegisterRow] = []
    for row_cells in rows_iter:
        row_dict = excel_row_to_dict_by_column_letter(row_cells, col_letter_to_header)
        item = s_row_to_register(row_dict)
        if item:
            result.append(item)
    wb.close()
    return s_columns, result


# ---------------------------------------------------------------------------
# Atlassian HTTP client
# ---------------------------------------------------------------------------


class AtlassianClient:
    def __init__(
        self,
        site_url: str,
        cloud_id: str,
        email: str,
        api_token: str,
        *,
        silent_mode: bool = True,
    ):
        self.site_url = site_url.rstrip("/")
        self.cloud_id = cloud_id
        self.silent_mode = silent_mode
        self.session = requests.Session()
        self.session.auth = (email, api_token)
        self.session.headers.update({"Accept": "application/json"})
        self.jira_base = f"https://api.atlassian.com/ex/jira/{cloud_id}/rest/api/3"
        self.jira_agile_base = f"https://api.atlassian.com/ex/jira/{cloud_id}/rest/agile/1.0"
        self.confluence_v2 = f"https://api.atlassian.com/ex/confluence/{cloud_id}/wiki/api/v2"

    def _jira_path(self, path: str, *, mutate: bool) -> str:
        """寫入類 API 附加 notifyUsers=false（同 jira_gui_cursor 寂靜同步）。"""
        if not mutate or not self.silent_mode:
            return path
        if "notifyUsers=" in path:
            return path
        sep = "&" if "?" in path else "?"
        return f"{path}{sep}notifyUsers=false"

    def jira_get(self, path: str, **params: Any) -> Any:
        r = self.session.get(f"{self.jira_base}{path}", params=params, timeout=60)
        if not r.ok:
            raise RuntimeError(f"Jira GET {path} failed: {r.status_code} {r.text}")
        return r.json() if r.text else {}

    def jira_search(
        self,
        jql: str,
        *,
        max_results: int = 100,
        fields: list[str] | None = None,
        next_page_token: str | None = None,
    ) -> Any:
        payload: dict[str, Any] = {
            "jql": jql,
            "maxResults": max_results,
            "fields": fields or ["summary", "status", "parent"],
        }
        if next_page_token:
            payload["nextPageToken"] = next_page_token
        # 搜尋不需 notifyUsers
        r = self.session.post(
            f"{self.jira_base}/search/jql", json=payload, timeout=60
        )
        if not r.ok:
            raise RuntimeError(f"Jira POST /search/jql failed: {r.status_code} {r.text}")
        return r.json() if r.text else {}

    def jira_post(self, path: str, payload: dict) -> Any:
        path = self._jira_path(path, mutate=True)
        r = self.session.post(
            f"{self.jira_base}{path}", json=payload, timeout=60
        )
        if not r.ok:
            raise RuntimeError(f"Jira POST {path} failed: {r.status_code} {r.text}")
        return r.json() if r.text else {}

    def jira_put(self, path: str, payload: dict) -> None:
        path = self._jira_path(path, mutate=True)
        r = self.session.put(
            f"{self.jira_base}{path}", json=payload, timeout=60
        )
        if not r.ok:
            raise RuntimeError(f"Jira PUT {path} failed: {r.status_code} {r.text}")

    def jira_delete(self, path: str) -> None:
        path = self._jira_path(path, mutate=True)
        r = self.session.delete(f"{self.jira_base}{path}", timeout=60)
        if not r.ok:
            raise RuntimeError(f"Jira DELETE {path} failed: {r.status_code} {r.text}")

    def jira_agile_put(self, path: str, payload: dict) -> None:
        r = self.session.put(
            f"{self.jira_agile_base}{path}", json=payload, timeout=60
        )
        if not r.ok:
            raise RuntimeError(
                f"Jira Agile PUT {path} failed: {r.status_code} {r.text}"
            )

    def confluence_get(self, path: str, **params: Any) -> Any:
        r = self.session.get(f"{self.confluence_v2}{path}", params=params, timeout=60)
        r.raise_for_status()
        return r.json()

    def confluence_put(self, path: str, payload: dict) -> Any:
        r = self.session.put(
            f"{self.confluence_v2}{path}", json=payload, timeout=60
        )
        if not r.ok:
            raise RuntimeError(f"Confluence PUT failed: {r.status_code} {r.text}")
        return r.json()


# ---------------------------------------------------------------------------
# Confluence
# ---------------------------------------------------------------------------


def _extract_jira_key_from_cell(cell_content: str) -> str | None:
    href_match = re.search(
        r'href="[^"]*/browse/(PMWC-\d+)"', cell_content, flags=re.IGNORECASE
    )
    if href_match:
        return href_match.group(1)
    text = re.sub(r"<[^>]+>", "", cell_content)
    m = CONFLUENCE_JIRA_LINK_RE.search(text) or JIRA_KEY_RE.search(text)
    return m.group(1) if m else None


def parse_jira_map_from_confluence_body(body: str) -> dict[str, str]:
    """從 C 表格 markdown/HTML 解析 ID -> PMWC key。

    優先讀 LINK 欄（最後一欄）；相容舊版將 JIRA 放在第 2 欄的頁面。
    """
    mapping: dict[str, str] = {}
    if "<table" in body.lower():
        for row_html in re.findall(r"<tr>(.*?)</tr>", body, flags=re.DOTALL | re.IGNORECASE):
            if re.search(r"<th\b", row_html, flags=re.IGNORECASE):
                continue
            cells = re.findall(r"<td[^>]*>(.*?)</td>", row_html, flags=re.DOTALL | re.IGNORECASE)
            if len(cells) < 2 or is_blank_table_row(cells):
                continue
            rid = normalize_register_id(re.sub(r"<[^>]+>", "", cells[0]))
            if not REGISTER_ID_RE.match(rid):
                continue
            key = _extract_jira_key_from_cell(cells[-1])
            if not key and len(cells) >= 2:
                key = _extract_jira_key_from_cell(cells[1])
            if key:
                mapping[rid] = key
        if mapping:
            return mapping

    for line in body.splitlines():
        if not line.strip().startswith("|"):
            continue
        cells = [strip_markdown_cell(c) for c in line.strip().strip("|").split("|")]
        if len(cells) < 2 or is_blank_table_row(cells):
            continue
        rid = normalize_register_id(cells[0])
        if not REGISTER_ID_RE.match(rid):
            continue
        key = _extract_jira_key_from_cell(cells[-1])
        if not key and len(cells) >= 2:
            key = _extract_jira_key_from_cell(cells[1])
        if key:
            mapping[rid] = key
    return mapping


def fetch_confluence_page(client: AtlassianClient, page_id: str) -> dict[str, Any]:
    return client.confluence_get(
        f"/pages/{page_id}",
        **{"body-format": "markdown"},
    )


def extract_confluence_body_text(page: dict[str, Any]) -> str:
    """從 Confluence v2 page JSON 取出正文（markdown / storage 等格式）。"""
    body = page.get("body") or {}
    if isinstance(body, str):
        return body
    for fmt in ("markdown", "storage", "view"):
        section = body.get(fmt)
        if isinstance(section, str):
            return section
        if isinstance(section, dict):
            value = section.get("value")
            if isinstance(value, str):
                return value
    value = body.get("value")
    return value if isinstance(value, str) else ""


def register_row_to_s_cell_values(
    row: RegisterRow, s_columns: list[str] | None = None
) -> list[str]:
    """依 S 表欄位順序輸出 C 表儲存格值（不含 LINK）。"""
    columns = s_columns if s_columns is not None else S_COLUMNS
    return [cell_value_for_s_column(row, col) for col in columns]


def register_row_to_c_cells(
    row: RegisterRow, s_columns: list[str] | None = None
) -> list[str]:
    """S 欄位 + LINK 欄文字值（LINK 為 PMWC key 或空）。"""
    cells = register_row_to_s_cell_values(row, s_columns)
    cells.append(row.jira_key if is_valid_jira_key(row.jira_key) else "")
    return cells


def jira_browse_url(key: str, site_url: str) -> str:
    return f"{site_url.rstrip('/')}/browse/{key}"


def jira_cell_html(key: str, site_url: str) -> str:
    url = jira_browse_url(key, site_url)
    return f'<a href="{escape_html(url)}">{escape_html(key)}</a>'


def jira_cell_markdown(key: str, site_url: str) -> str:
    url = jira_browse_url(key, site_url)
    return f"[{key}]({url})"


def build_confluence_html_table(
    rows: list[RegisterRow], site_url: str, s_columns: list[str] | None = None
) -> str:
    """以 Confluence storage HTML 輸出表格：S 欄位順序 + 最後 LINK 欄。"""
    columns = c_columns_for(s_columns) if s_columns is not None else C_COLUMNS
    data_columns = s_columns if s_columns is not None else S_COLUMNS
    parts = ['<table data-layout="full-width"><tbody>']
    header_cells = "".join(
        f"<th><p><strong>{escape_html(col)}</strong></p></th>" for col in columns
    )
    parts.append(f"<tr>{header_cells}</tr>")
    for row in rows:
        row_cells = [
            f"<td><p>{escape_html(value)}</p></td>"
            for value in register_row_to_s_cell_values(row, data_columns)
        ]
        link_inner = (
            jira_cell_html(row.jira_key, site_url)
            if is_valid_jira_key(row.jira_key)
            else ""
        )
        row_cells.append(f"<td><p>{link_inner}</p></td>")
        parts.append(f"<tr>{''.join(row_cells)}</tr>")
    parts.append("</tbody></table>")
    return "".join(parts)


# CONFIG_YAML 若在 Windows 以錯誤編碼寫入 secret，中文會變成 ?；以腳本內 UTF-8 預設救回。
def _title_looks_corrupted(text: str) -> bool:
    """偵測編碼損壞（中文變 ? 或 U+FFFD）。合法標題不應含 ASCII ?。"""
    t = (text or "").strip()
    if not t:
        return True
    return "?" in t or "\ufffd" in t


def resolve_sharepoint_source_link(cfg: dict[str, Any]) -> tuple[str, str]:
    """回傳 Confluence 頁首 S 表連結 (url, title)。"""
    conf = cfg.get("confluence", {})
    sp = cfg.get("sharepoint", {})
    url = (
        conf.get("source_link_url")
        or sp.get("view_url")
        or sp.get("download_url")
        or ""
    ).strip()
    title = (conf.get("source_link_title") or "").strip()
    if _title_looks_corrupted(title):
        title = DEFAULT_SOURCE_LINK_TITLE
    return url, title


def build_confluence_source_link_html(url: str, title: str) -> str:
    """Confluence 頁首：僅 S 表來源連結（不再渲染「立即同步」按鈕）。"""
    if not url:
        return ""
    safe_title = title.strip() if not _title_looks_corrupted(title) else DEFAULT_SOURCE_LINK_TITLE
    return (
        f"<p><strong>S 表格：</strong>"
        f'<a href="{escape_html(url)}">{escape_html(safe_title)}</a></p>'
    )


def build_confluence_page_html(
    rows: list[RegisterRow],
    site_url: str,
    s_columns: list[str] | None = None,
    *,
    source_link_url: str = "",
    source_link_title: str = "",
) -> str:
    """Confluence 頁面正文：S 表連結 + 資料表格。"""
    prefix = build_confluence_source_link_html(source_link_url, source_link_title)
    table = build_confluence_html_table(rows, site_url, s_columns)
    return prefix + table


def build_confluence_markdown(
    rows: list[RegisterRow], site_url: str, s_columns: list[str] | None = None
) -> str:
    """保留 markdown 輸出供除錯；正式更新請用 build_confluence_html_table。"""
    columns = c_columns_for(s_columns) if s_columns is not None else C_COLUMNS
    data_columns = s_columns if s_columns is not None else S_COLUMNS
    header = "| " + " | ".join(columns) + " |"
    sep = "| " + " | ".join(["---"] * len(columns)) + " |"
    lines = [header, sep]
    for row in rows:
        cells = register_row_to_s_cell_values(row, data_columns)
        link_cell = (
            jira_cell_markdown(row.jira_key, site_url)
            if is_valid_jira_key(row.jira_key)
            else ""
        )
        cells.append(link_cell)
        escaped = [cell.replace("|", "\\|").replace("\n", " ") for cell in cells]
        lines.append("| " + " | ".join(escaped) + " |")
    return "\n".join(lines) + "\n"


def update_confluence_page(
    client: AtlassianClient,
    page_id: str,
    title: str,
    table_html: str,
    version_message: str,
    dry_run: bool,
) -> bool:
    page = fetch_confluence_page(client, page_id)
    version = page["version"]["number"]
    # minorEdit=True：比照 D:\MEGA\Jira 其他 Confluence 腳本，降低追蹤者通知
    payload = {
        "id": page_id,
        "status": "current",
        "title": title,
        "body": {"representation": "storage", "value": table_html},
        "version": {
            "number": version + 1,
            "message": version_message,
            "minorEdit": True,
        },
    }
    if dry_run:
        print(
            f"[dry-run] Confluence page {page_id} would update to v{version + 1} "
            f"({len(table_html)} bytes HTML, {table_html.count('<tr>') - 1} data rows)"
        )
        return False
    client.confluence_put(f"/pages/{page_id}", payload)
    return True


# ---------------------------------------------------------------------------
# Jira
# ---------------------------------------------------------------------------


def build_timeline_fields(
    row: RegisterRow, sync_date: str, cfg: dict[str, Any]
) -> dict[str, str]:
    """設定 Jira Timeline：Opened → 起始日；Target close → 截止日。

    嚴格規則（空白／無法解析 → 不出現在回傳 dict，Jira 更新不會動該欄）：
    - 儲存格空白／空字串 → 省略對應欄位，**不變更** Jira 既有值
    - 有內容但無法解析為日期 → 忽略，**不變更** Jira
    - 成功解析 → 寫入該欄（Start＝Opened，Due＝Target close）
    - 兩者皆成功解析且 due < start → 僅此時將 due 年 +1
    - 絕不回填 sync_date、不以 null/None 清空欄位

    sync_date 保留供呼叫端相容；不參與日期推導。
    """
    del sync_date  # 保留參數簽名；日期僅依成功解析的 Opened / Target close
    opened = parse_flexible_date(row.opened)
    due = parse_flexible_date(row.target_close)
    if not opened and not due:
        return {}

    jira_cfg = cfg.get("jira", {})
    start_field = jira_cfg.get("start_date_field", "customfield_10015")
    due_field = jira_cfg.get("due_date_field", "duedate")
    out: dict[str, str] = {}

    if opened:
        out[start_field] = opened

    if due:
        # year+1 僅在「本 run 兩者皆自表格解析成功」時套用
        if opened:
            due_dt = datetime.strptime(due, "%Y-%m-%d")
            start_dt = datetime.strptime(opened, "%Y-%m-%d")
            if due_dt < start_dt:
                due_dt = due_dt.replace(year=start_dt.year + 1)
                due = due_dt.strftime("%Y-%m-%d")
        out[due_field] = due

    return out


def row_has_timeline(row: RegisterRow) -> bool:
    """Opened 或 Target close 有可解析日期（會寫入 Start／Due）即視為有時間軸。"""
    return bool(
        parse_flexible_date(row.opened) or parse_flexible_date(row.target_close)
    )


def timeline_rank_group(row: RegisterRow) -> int:
    """時間軸排序群組：0=有時間軸 → 1=進行中(僅排序用) → 2=其他開放 → 3=完成。

    已完成永遠置底（即使曾有 Target close）。
    Jira 狀態同步僅會在 Status 含 Done/Closed/Completed 時寫入「完成」；
    此處「進行中」僅參考原文關鍵字，不觸發 Jira transition。
    """
    src = jira_status_source(row)
    if map_status_to_jira(src) == JIRA_DONE_STATUS:
        return 3
    if row_has_timeline(row):
        return 0
    lowered = (src or "").casefold()
    if "in progress" in lowered or "in-progress" in lowered:
        return 1
    return 2


def sort_rows_for_jira_rank(rows: list[RegisterRow]) -> list[RegisterRow]:
    """依群組 + register_id 排序，供 Jira Rank（越小越上面）。"""
    return sorted(
        rows,
        key=lambda r: (timeline_rank_group(r), r.register_id.upper()),
    )


def apply_jira_rank_order(
    client: AtlassianClient,
    ordered_keys: list[str],
    dry_run: bool,
) -> None:
    """將 issues 依 ordered_keys 順序排 Rank（前者在上）。Agile API 單次最多 50 筆。"""
    keys = [k for k in ordered_keys if is_valid_jira_key(k)]
    if len(keys) < 2:
        return
    if dry_run:
        print(f"[dry-run] Rank {len(keys)} issues: {keys[0]} ... {keys[-1]}")
        return
    head = keys[0]
    rest = keys[1:]
    # 先把第一筆排到第二筆前面，確保 head 在相對頂部
    client.jira_agile_put(
        "/issue/rank",
        {"issues": [head], "rankBeforeIssue": rest[0]},
    )
    for i in range(0, len(rest), 50):
        chunk = rest[i : i + 50]
        anchor = head if i == 0 else rest[i - 1]
        client.jira_agile_put(
            "/issue/rank",
            {"issues": chunk, "rankAfterIssue": anchor},
        )


def reorder_jira_timeline_view(
    client: AtlassianClient,
    rows: list[RegisterRow],
    report: SyncReport,
    dry_run: bool,
) -> None:
    """依 Priority (Epic 群組) 分組後重排 Rank：有時間軸 → 進行中 → 其他 → 完成。"""
    by_group: dict[str, list[RegisterRow]] = {}
    for row in rows:
        if not is_valid_jira_key(row.jira_key):
            continue
        by_group.setdefault(jira_epic_priority(row), []).append(row)

    for label, group in sorted(by_group.items()):
        ordered = sort_rows_for_jira_rank(group)
        ordered_keys = [r.jira_key for r in ordered]
        try:
            apply_jira_rank_order(client, ordered_keys, dry_run)
            report.ranked_issues.extend(ordered_keys)
            print(
                f"Jira Rank ({label}): "
                f"時間軸={sum(1 for r in ordered if timeline_rank_group(r) == 0)}, "
                f"進行中={sum(1 for r in ordered if timeline_rank_group(r) == 1)}, "
                f"其他={sum(1 for r in ordered if timeline_rank_group(r) == 2)}, "
                f"完成={sum(1 for r in ordered if timeline_rank_group(r) == 3)}"
            )
        except Exception as exc:  # noqa: BLE001
            report.errors.append(f"Jira Rank ({label}): {exc}")


def resolve_default_assignee_account_id(
    client: AtlassianClient, cfg: dict[str, Any]
) -> str:
    """取得預設受託人 accountId（設定檔優先，否則以 displayName 搜尋）。"""
    jira_cfg = cfg.get("jira", {})
    account_id = str(jira_cfg.get("default_assignee_account_id") or "").strip()
    if account_id:
        return account_id
    query = str(jira_cfg.get("default_assignee_query") or "shannon").strip()
    if not query:
        return ""
    users = client.jira_get("/user/search", query=query)
    if not isinstance(users, list):
        return ""
    lowered = query.lower()
    for user in users:
        if not user.get("active", True):
            continue
        display = str(user.get("displayName") or "").lower()
        email = str(user.get("emailAddress") or "").lower()
        if lowered in display or lowered in email:
            return str(user.get("accountId") or "")
    if users:
        return str(users[0].get("accountId") or "")
    return ""


def ensure_default_assignee(
    client: AtlassianClient,
    issue_key: str,
    account_id: str,
    report: SyncReport,
    dry_run: bool,
) -> None:
    """受託人為空時指派預設帳號（Shannon）。"""
    if not account_id or not is_valid_jira_key(issue_key):
        return
    if dry_run:
        print(f"[dry-run] Assign {issue_key} -> {account_id}")
        return
    try:
        data = client.jira_get(f"/issue/{issue_key}", fields="assignee")
        assignee = (data.get("fields") or {}).get("assignee")
        if assignee and assignee.get("accountId"):
            return
        client.jira_put(f"/issue/{issue_key}/assignee", {"accountId": account_id})
        report.assigned_issues.append(issue_key)
    except Exception as exc:  # noqa: BLE001
        report.errors.append(f"指派受託人 {issue_key}: {exc}")


def assign_unassigned_register_issues(
    client: AtlassianClient,
    rows: list[RegisterRow],
    cfg: dict[str, Any],
    report: SyncReport,
    dry_run: bool,
) -> None:
    """同步後：未指派的 register 任務一律指派給預設受託人。"""
    account_id = resolve_default_assignee_account_id(client, cfg)
    if not account_id:
        report.errors.append("找不到預設受託人（請設 jira.default_assignee_account_id）")
        return
    print(f"預設受託人 accountId: {account_id}")
    for row in rows:
        if is_valid_jira_key(row.jira_key):
            ensure_default_assignee(
                client, row.jira_key, account_id, report, dry_run
            )


def find_duplicate_register_ids(rows: list[RegisterRow]) -> set[str]:
    """回傳在 rows 中出現超過一次的 register_id（大小寫已正規化）。"""
    counts: dict[str, int] = {}
    for row in rows:
        rid = (row.register_id or "").strip()
        if not rid:
            continue
        counts[rid] = counts.get(rid, 0) + 1
    return {rid for rid, n in counts.items() if n > 1}


def jira_summary(
    register_id: str, title: str, *, duplicate: bool = False
) -> str:
    """Jira 任務名稱 = ID_Title；S 表 ID 重複時最前方加「ID重複」。"""
    safe_title = title.replace("\n", " ").strip()
    base = f"{register_id}_{safe_title}"
    if duplicate:
        base = f"{DUPLICATE_ID_SUMMARY_PREFIX} {base}"
    return base[:255]


def text_to_adf(text: str) -> dict[str, Any]:
    """將純文字轉為 Jira Atlassian Document Format (ADF)。"""
    paragraphs: list[dict[str, Any]] = []
    for block in text.splitlines():
        block = block.strip()
        if not block:
            continue
        paragraphs.append(
            {
                "type": "paragraph",
                "content": [{"type": "text", "text": block}],
            }
        )
    if not paragraphs:
        paragraphs.append(
            {
                "type": "paragraph",
                "content": [{"type": "text", "text": ""}],
            }
        )
    return {"type": "doc", "version": 1, "content": paragraphs}


def build_jira_description(row: RegisterRow) -> dict[str, Any]:
    lines = [f"Register ID: {row.register_id}"]
    if row.title:
        lines.append(f"Title: {row.title}")
    if row.description:
        lines.append(row.description)
    return text_to_adf("\n".join(lines))


def build_confluence_page_url(cfg: dict[str, Any], page: dict[str, Any] | None = None) -> str:
    """組出 mail_checking 等 Confluence 頁面 URL（供 Jira Web 連結）。"""
    if page:
        links = page.get("_links") or {}
        base = str(links.get("base", "")).rstrip("/")
        webui = links.get("webui", "")
        if base and webui:
            return f"{base}{webui}".split("#")[0]
        web_url = page.get("webUrl")
        if isinstance(web_url, str) and web_url.startswith("http"):
            return web_url.split("#")[0]

    site = cfg["atlassian"]["site_url"].rstrip("/")
    conf = cfg["confluence"]
    page_id = conf["page_id"]
    title = conf.get("page_title", "mail_checking")
    space = conf.get("space_key", "projectmod")
    return f"{site}/wiki/spaces/{space}/pages/{page_id}/{title}"


def ensure_confluence_web_link(
    client: AtlassianClient,
    issue_key: str,
    page_url: str,
    link_title: str,
    report: SyncReport,
    dry_run: bool,
    *,
    page_id: str = "",
) -> None:
    """在 Jira issue 新增指向 Confluence C 表的 Web 連結（remotelink）。"""
    if dry_run:
        print(f"[dry-run] Web 連結 {issue_key} -> {link_title}: {page_url}")
        return
    try:
        existing = client.jira_get(f"/issue/{issue_key}/remotelink")
        links = existing if isinstance(existing, list) else []
        stale_ids: list[int | str] = []
        has_correct = False
        for link in links:
            obj = link.get("object") or {}
            url = obj.get("url") or ""
            title = obj.get("title") or ""
            if url == page_url:
                if title == link_title:
                    has_correct = True
                else:
                    link_id = link.get("id")
                    if link_id is not None:
                        stale_ids.append(link_id)
                continue
            if title == link_title or (page_id and page_id in url):
                link_id = link.get("id")
                if link_id is not None:
                    stale_ids.append(link_id)
        for link_id in stale_ids:
            client.jira_delete(f"/issue/{issue_key}/remotelink/{link_id}")
        if has_correct:
            return
        client.jira_post(
            f"/issue/{issue_key}/remotelink",
            {"object": {"url": page_url, "title": link_title}},
        )
        report.web_links_added.append(issue_key)
    except Exception as exc:  # noqa: BLE001
        report.errors.append(f"Web 連結 {issue_key}: {exc}")


def strip_duplicate_id_summary_prefix(summary: str) -> str:
    """移除 summary 最前方的「ID重複」標記（含空白或底線分隔）。"""
    text = (summary or "").strip()
    markers = (
        f"{DUPLICATE_ID_SUMMARY_PREFIX} ",
        f"{DUPLICATE_ID_SUMMARY_PREFIX}_",
        DUPLICATE_ID_SUMMARY_PREFIX,
    )
    for marker in markers:
        if text.startswith(marker):
            return text[len(marker) :].lstrip(" _")
    return text


def parse_register_id_from_summary(summary: str) -> str | None:
    text = strip_duplicate_id_summary_prefix(summary)
    if "_" not in text:
        return None
    prefix = text.split("_", 1)[0].strip().upper()
    if REGISTER_ID_RE.match(prefix):
        return prefix
    return None


def _paginate_jira_search(
    client: AtlassianClient,
    jql: str,
    *,
    fields: list[str],
    max_results: int = 100,
) -> list[dict[str, Any]]:
    """分頁蒐集 JQL 搜尋結果。"""
    collected: list[dict[str, Any]] = []
    next_token: str | None = None
    while True:
        data = client.jira_search(
            jql,
            max_results=max_results,
            fields=fields,
            next_page_token=next_token,
        )
        issues = data.get("issues", [])
        collected.extend(issues)
        if data.get("isLast", True) or not issues:
            break
        next_token = data.get("nextPageToken")
        if not next_token:
            break
    return collected


def load_jira_register_index(
    client: AtlassianClient, project_key: str, task_type: str
) -> dict[str, dict[str, Any]]:
    """register_id -> {key, status, parent}"""
    index: dict[str, dict[str, Any]] = {}
    issues: list[dict[str, Any]] = []
    for type_name in jql_issuetype_name_candidates(task_type):
        jql = f"project = {project_key} AND {jql_issuetype_equals(type_name)}"
        issues = _paginate_jira_search(
            client, jql, fields=["summary", "status", "parent"]
        )
        if issues:
            break
    for issue in issues:
        rid = parse_register_id_from_summary(issue["fields"]["summary"])
        if rid:
            parent = issue["fields"].get("parent") or {}
            index[rid] = {
                "key": issue["key"],
                "status": issue["fields"]["status"]["name"],
                "parent_key": parent.get("key", ""),
            }
    return index


def load_epic_map(
    client: AtlassianClient, project_key: str, epic_type: str
) -> dict[str, str]:
    """priority label (P1) -> epic key"""
    mapping: dict[str, str] = {}
    issues: list[dict[str, Any]] = []
    for type_name in jql_issuetype_name_candidates(epic_type):
        jql = f"project = {project_key} AND {jql_issuetype_equals(type_name)}"
        issues = _paginate_jira_search(client, jql, fields=["summary"])
        if issues:
            break
    for issue in issues:
        summary = issue["fields"]["summary"].strip()
        mapping[summary] = issue["key"]
    return mapping


def find_jira_key_by_register_id(
    client: AtlassianClient,
    project_key: str,
    task_type: str,
    register_id: str,
) -> str:
    """以 summary 前綴搜尋 Jira 任務（補 issue_index 遺漏）。"""
    for type_name in jql_issuetype_name_candidates(task_type):
        jql = (
            f"project = {project_key} AND {jql_issuetype_equals(type_name)} "
            f'AND summary ~ "{register_id}"'
        )
        try:
            data = client.jira_search(jql, max_results=20, fields=["summary"])
        except Exception:
            continue
        for issue in data.get("issues", []):
            rid = parse_register_id_from_summary(issue["fields"]["summary"])
            if rid == register_id:
                return issue["key"]
    return ""


def ensure_epic(
    client: AtlassianClient,
    project_key: str,
    epic_type: str,
    priority: str,
    epic_map: dict[str, str],
    report: SyncReport,
    dry_run: bool,
) -> str:
    if priority in epic_map:
        return epic_map[priority]
    if dry_run:
        fake = f"PMWC-NEW-{priority}"
        print(f"[dry-run] Would create epic: {priority} -> {fake}")
        epic_map[priority] = fake
        report.created_epics.append(priority)
        return fake
    created = client.jira_post(
        "/issue",
        {
            "fields": {
                "project": {"key": project_key},
                "issuetype": {"name": epic_type},
                "summary": priority,
            }
        },
    )
    key = created["key"]
    epic_map[priority] = key
    report.created_epics.append(f"{priority} ({key})")
    return key


def get_transition_id(client: AtlassianClient, issue_key: str, target_status: str) -> str | None:
    target = normalize_jira_status_name(target_status)
    if not target:
        return None
    data = client.jira_get(f"/issue/{issue_key}/transitions")
    for tr in data.get("transitions", []):
        to_name = normalize_jira_status_name(tr.get("to", {}).get("name", ""))
        if to_name == target:
            return tr["id"]
        if to_name.casefold() == target.casefold():
            return tr["id"]
    return None


def transition_issue(
    client: AtlassianClient,
    issue_key: str,
    target_status: str,
    dry_run: bool,
) -> None:
    if dry_run:
        print(f"[dry-run] Transition {issue_key} -> {target_status}")
        return
    tr_id = get_transition_id(client, issue_key, target_status)
    if not tr_id:
        raise RuntimeError(
            f"無法將 {issue_key} 轉換到狀態 {target_status!r}（找不到 transition）"
        )
    client.jira_post(f"/issue/{issue_key}/transitions", {"transition": {"id": tr_id}})


def delete_jira_issue(
    client: AtlassianClient, issue_key: str, dry_run: bool
) -> None:
    """刪除 C 表已不存在的 Jira 任務。"""
    if dry_run:
        print(f"[dry-run] Delete Jira issue {issue_key}")
        return
    client.jira_delete(f"/issue/{issue_key}?deleteSubtasks=false")


def sync_jira_for_rows(
    client: AtlassianClient,
    rows: list[RegisterRow],
    cfg: dict[str, Any],
    existing_jira_map: dict[str, str],
    report: SyncReport,
    dry_run: bool,
    sync_date: str,
    confluence_page_url: str,
    web_link_title: str,
) -> dict[str, str]:
    project = cfg["jira"]["project_key"]
    epic_type = resolve_issuetype_for_api(
        cfg["jira"]["epic_issue_type"], role="epic"
    )
    task_type = resolve_issuetype_for_api(
        cfg["jira"]["task_issue_type"], role="task"
    )

    epic_map = load_epic_map(client, project, epic_type)
    issue_index = load_jira_register_index(client, project, task_type)

    current_ids = {r.register_id for r in rows}
    id_to_key: dict[str, str] = dict(existing_jira_map)
    duplicate_ids = find_duplicate_register_ids(rows)
    if duplicate_ids:
        dup_list = ", ".join(sorted(duplicate_ids))
        print(f"警告: S 表重複 Register ID（Jira 標題加前綴「{DUPLICATE_ID_SUMMARY_PREFIX}」）: {dup_list}")

    # 刪除 C 表已不存在的項目（Jira 與 C 表保持一致）
    for rid, info in issue_index.items():
        if rid not in current_ids:
            if dry_run:
                print(f"[dry-run] Delete removed register {rid} ({info['key']})")
                report.deleted_issues.append(info["key"])
            else:
                try:
                    delete_jira_issue(client, info["key"], dry_run=False)
                    report.deleted_issues.append(info["key"])
                except Exception as exc:  # noqa: BLE001
                    report.errors.append(f"刪除 {info['key']}: {exc}")

    for row in rows:
        priority = jira_epic_priority(row)
        epic_key = ensure_epic(
            client, project, epic_type, priority, epic_map, report, dry_run
        )
        target_status = map_status_to_jira(jira_status_source(row))
        summary = jira_summary(
            row.register_id,
            row.title,
            duplicate=row.register_id in duplicate_ids,
        )
        description = build_jira_description(row)

        existing = issue_index.get(row.register_id)
        known_key = id_to_key.get(row.register_id) or (
            existing["key"] if existing else ""
        )
        if not known_key:
            known_key = find_jira_key_by_register_id(
                client, project, task_type, row.register_id
            )

        timeline_fields = build_timeline_fields(row, sync_date, cfg)

        if not known_key:
            if dry_run:
                fake_key = f"PMWC-NEW-{row.register_id}"
                timeline_msg = (
                    f", timeline {timeline_fields}" if timeline_fields else ""
                )
                print(
                    f"[dry-run] Create {row.register_id} under {epic_key} "
                    f"-> {fake_key}{timeline_msg}"
                )
                id_to_key[row.register_id] = fake_key
                report.created_issues.append(fake_key)
                row.jira_key = fake_key
                ensure_confluence_web_link(
                    client,
                    fake_key,
                    confluence_page_url,
                    web_link_title,
                    report,
                    dry_run,
                    page_id=str(cfg["confluence"]["page_id"]),
                )
                continue
            create_fields: dict[str, Any] = {
                "project": {"key": project},
                "issuetype": {"name": task_type},
                "summary": summary,
                "description": description,
                "parent": {"key": epic_key},
            }
            create_fields.update(timeline_fields)
            try:
                created = client.jira_post("/issue", {"fields": create_fields})
            except Exception as exc:  # noqa: BLE001
                report.errors.append(f"建立 {row.register_id}: {exc}")
                continue
            key = created["key"]
            id_to_key[row.register_id] = key
            row.jira_key = key
            report.created_issues.append(key)
            issue_index[row.register_id] = {
                "key": key,
                "status": DEFAULT_JIRA_STATUS,
                "parent_key": epic_key,
            }
            if target_status:
                try:
                    transition_issue(client, key, target_status, dry_run=False)
                    issue_index[row.register_id]["status"] = target_status
                except Exception as exc:  # noqa: BLE001
                    report.errors.append(f"設定狀態 {key}: {exc}")
            ensure_confluence_web_link(
                client,
                key,
                confluence_page_url,
                web_link_title,
                report,
                dry_run,
                page_id=str(cfg["confluence"]["page_id"]),
            )
            continue

        if not is_valid_jira_key(known_key):
            known_key = find_jira_key_by_register_id(
                client, project, task_type, row.register_id
            )
        if not known_key:
            report.errors.append(f"{row.register_id}: 找不到有效 JIRA key")
            continue

        row.jira_key = known_key
        id_to_key[row.register_id] = known_key
        fields: dict[str, Any] = {
            "summary": summary,
            "description": description,
            "parent": {"key": epic_key},
        }
        fields.update(timeline_fields)
        if dry_run:
            timeline_msg = f", timeline={timeline_fields}" if timeline_fields else ""
            status_msg = (
                f", status->{target_status}"
                if target_status
                else ", status=(不變)"
            )
            print(
                f"[dry-run] Update {known_key}: parent={epic_key}"
                f"{status_msg}{timeline_msg}"
            )
        else:
            try:
                client.jira_put(f"/issue/{known_key}", {"fields": fields})
                report.updated_issues.append(known_key)
            except Exception as exc:  # noqa: BLE001
                report.errors.append(f"更新 {known_key}: {exc}")

            if target_status:
                current_status = issue_index.get(row.register_id, {}).get("status", "")
                if current_status != target_status:
                    try:
                        transition_issue(client, known_key, target_status, dry_run=False)
                    except Exception as exc:  # noqa: BLE001
                        report.errors.append(f"轉換狀態 {known_key}: {exc}")

        ensure_confluence_web_link(
            client,
            known_key,
            confluence_page_url,
            web_link_title,
            report,
            dry_run,
            page_id=str(cfg["confluence"]["page_id"]),
        )

    return id_to_key


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def run_sync(config_path: Path, dry_run: bool) -> SyncReport:
    cfg = load_config(config_path)
    report = SyncReport()

    cache_dir = SCRIPT_DIR / ".register_cache"
    snapshot_path = cache_dir / "last_snapshot.json"
    log_dir = SCRIPT_DIR / "logs"

    old_snapshot, snapshot_saved_at = load_last_snapshot(snapshot_path)

    # 下載前讀取上次檔名，供郵件主旨顯示改名過渡（下載會覆寫 register_original_name.txt）
    s_previous_filename = read_cached_register_original_name(cache_dir)

    xlsx, s_original_filename = download_sharepoint_excel(
        cfg["sharepoint"]["download_url"], cache_dir
    )
    print(f"已下載 S 表格: {xlsx} ({xlsx.stat().st_size} bytes)")
    subject_preview = resolve_register_subject_token(
        s_original_filename,
        cfg,
        previous_filename=s_previous_filename,
    )
    if subject_preview:
        print(f"郵件主旨括號: [{subject_preview}]")

    s_columns, s_rows = load_register_rows_from_excel(xlsx, cfg["sharepoint"]["sheet_name"])
    print(f"S 表格欄位 ({len(s_columns)}): {', '.join(s_columns)}")
    print(f"S 表格有效項目: {len(s_rows)}")
    if not s_rows:
        raise SystemExit("S 表格沒有可同步的項目。")

    register_diff = compute_register_diff(
        old_snapshot, s_rows, snapshot_saved_at=snapshot_saved_at
    )

    client = AtlassianClient(
        cfg["atlassian"]["site_url"],
        cfg["atlassian"]["cloud_id"],
        cfg["atlassian"]["email"],
        cfg["_api_token"],
        silent_mode=bool(cfg.get("silent_mode", True)),
    )
    if client.silent_mode:
        print("寂靜寫入：Jira notifyUsers=false、Confluence minorEdit=true")

    page_id = str(cfg["confluence"]["page_id"])
    page = fetch_confluence_page(client, page_id)
    body = extract_confluence_body_text(page)
    existing_jira_map = parse_jira_map_from_confluence_body(body)
    print(f"從 C 表格讀到 {len(existing_jira_map)} 筆 JIRA 對照")

    confluence_page_url = build_confluence_page_url(cfg, page)
    web_link_title = (
        cfg["confluence"].get("web_link_title")
        or cfg["confluence"].get("page_title", "mail_checking")
    )
    print(f"Jira Web 連結目標: {web_link_title} ({confluence_page_url})")

    sync_date = datetime.now().strftime("%Y-%m-%d")
    print(f"同步執行日: {sync_date}")

    id_to_key = sync_jira_for_rows(
        client,
        s_rows,
        cfg,
        existing_jira_map,
        report,
        dry_run,
        sync_date,
        confluence_page_url,
        web_link_title,
    )

    site_url = cfg["atlassian"]["site_url"]
    source_link_url, source_link_title = resolve_sharepoint_source_link(cfg)
    if source_link_url:
        print(f"S 表連結: {source_link_title} ({source_link_url[:60]}...)")
    for row in s_rows:
        key = row.jira_key or id_to_key.get(row.register_id, "")
        if not is_valid_jira_key(key):
            key = find_jira_key_by_register_id(
                client,
                cfg["jira"]["project_key"],
                cfg["jira"]["task_issue_type"],
                row.register_id,
            )
        if is_valid_jira_key(key):
            row.jira_key = key
            id_to_key[row.register_id] = key

    reorder_jira_timeline_view(client, s_rows, report, dry_run)
    assign_unassigned_register_issues(client, s_rows, cfg, report, dry_run)

    missing_jira = [
        row.register_id
        for row in s_rows
        if not is_valid_jira_key(row.jira_key)
    ]
    if missing_jira and not dry_run:
        msg = f"下列項目缺少 JIRA key: {', '.join(missing_jira)}"
        report.errors.append(msg)
        print(f"警告: {msg}")
        print("Confluence 未更新（需先補齊 JIRA）。")
        deliver_diff_report(
            cfg,
            log_dir,
            register_diff,
            report,
            dry_run=dry_run,
            s_link_url=source_link_url,
            s_link_title=source_link_title,
            confluence_url=confluence_page_url,
            s_original_filename=s_original_filename,
            s_previous_filename=s_previous_filename,
        )
        return report

    page_body = build_confluence_page_html(
        s_rows,
        site_url,
        s_columns,
        source_link_url=source_link_url,
        source_link_title=source_link_title,
    )
    print(f"C 表格將寫入 {len(s_rows)} 列 × {len(c_columns_for(s_columns))} 欄（無空白列）")
    if source_link_url:
        print("C 表頂部已加入 S 表格連結")
    report.confluence_updated = update_confluence_page(
        client,
        page_id,
        cfg["confluence"].get("page_title", page.get("title", "mail_checking")),
        page_body,
        cfg["confluence"].get("version_message", "Synced from SharePoint register"),
        dry_run,
    )

    if not dry_run:
        save_last_snapshot(snapshot_path, s_rows)
        print(f"已更新快照: {snapshot_path}")

    deliver_diff_report(
        cfg,
        log_dir,
        register_diff,
        report,
        dry_run=dry_run,
        s_link_url=source_link_url,
        s_link_title=source_link_title,
        confluence_url=confluence_page_url,
        s_original_filename=s_original_filename,
        s_previous_filename=s_previous_filename,
    )

    return report


def print_report(report: SyncReport) -> None:
    print("\n========== 同步結果 ==========")
    if report.created_epics:
        print(f"新建 Epic: {', '.join(report.created_epics)}")
    if report.created_issues:
        print(f"新建任務: {', '.join(report.created_issues)}")
    if report.updated_issues:
        print(f"更新任務: {', '.join(report.updated_issues)}")
    if report.deleted_issues:
        print(f"刪除任務: {', '.join(report.deleted_issues)}")
    if report.web_links_added:
        print(f"新增 Web 連結: {', '.join(report.web_links_added)}")
    if report.ranked_issues:
        print(f"重排 Rank: {len(report.ranked_issues)} 筆")
    if report.assigned_issues:
        print(f"指派受託人: {', '.join(report.assigned_issues)}")
    print(f"Confluence 已更新: {'是' if report.confluence_updated else '否（dry-run 或未變更）'}")
    if report.diff_log_path:
        print(f"變更紀錄: {report.diff_log_path}")
    if report.errors:
        print("\n錯誤:")
        for err in report.errors:
            print(f"  - {err}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="SharePoint Register -> Confluence -> Jira 同步"
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=SCRIPT_DIR / "config.yaml",
        help="設定檔路徑（預設 ./config.yaml）",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="預覽變更，不寫入 Confluence / Jira",
    )
    args = parser.parse_args()

    if not args.config.exists():
        print(f"找不到設定檔 {args.config}")
        print("請複製 config.example.yaml 為 config.yaml 並填入帳號資訊。")
        sys.exit(1)

    try:
        report = run_sync(args.config, dry_run=args.dry_run)
        print_report(report)
    except Exception as exc:  # noqa: BLE001
        print(f"同步失敗: {exc}", file=sys.stderr)
        sys.exit(1)

    # GitHub Actions：寄信失敗必須讓 workflow 變紅，避免漏通知
    if is_github_actions() and any(
        "寄送差異郵件" in err for err in report.errors
    ):
        sys.exit(1)


if __name__ == "__main__":
    main()
